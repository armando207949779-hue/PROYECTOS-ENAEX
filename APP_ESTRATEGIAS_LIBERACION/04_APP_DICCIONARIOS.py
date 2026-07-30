# ============================================================
# 04_APP_DICCIONARIOS
# APP_ESTRATEGIAS_LIBERACION
#
# Consulta los dos diccionarios cargados como archivos independientes:
# - Diccionario CECO-Plantas: CECO | Planta | Centro
# - Diccionario Usuarios-Cargos: Correo | Cargo
#
# No utiliza el Excel maestro antiguo ni Dic_Rangos.
# ============================================================

from __future__ import annotations

import base64
import html
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from textwrap import dedent
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURACIÓN
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
CHILE_TZ = ZoneInfo("America/Santiago")

LOGO_CANDIDATES = [
    PROJECT_DIR / "assets" / "logo.svg",
    BASE_DIR / "assets" / "logo.svg",
    PROJECT_DIR / "assets" / "logo.png",
    BASE_DIR / "assets" / "logo.png",
    PROJECT_DIR / "assets" / "logo.jpg",
    BASE_DIR / "assets" / "logo.jpg",
    PROJECT_DIR / "assets" / "logo.jpeg",
    BASE_DIR / "assets" / "logo.jpeg",
]

SESSION_DATA_KEY = "flujo_liberacion_data"
SESSION_FILE_KEY = "flujo_liberacion_file_name"
SESSION_SOURCE_FILES_KEY = "flujo_liberacion_source_files_v05"

DICTIONARY_CONFIG = {
    "cecos": {
        "titulo": "CECO Y PLANTAS",
        "icono": "🏭",
        "session_key": "dic_ceco",
        "file_role": "dic_ceco",
        "columns": ["CECO", "Planta", "Centro"],
        "description": "Catálogo de centros de costo, plantas y centros.",
    },
    "usuarios": {
        "titulo": "USUARIOS Y CARGOS",
        "icono": "👥",
        "session_key": "dic_users",
        "file_role": "dic_users",
        "columns": ["Correo", "Cargo"],
        "description": "Catálogo de correos y cargos asociados.",
    },
}



# ============================================================
# ESTILOS
# ============================================================

def compact_html(value: str) -> str:
    return re.sub(r">\s+<", "><", dedent(value).strip())


def aplicar_estilos() -> None:
    st.markdown(
        """
        <style>
            .stMainBlockContainer,
            .block-container {
                padding-top: 6.5rem !important;
                padding-bottom: 3rem !important;
            }

            .app-logo {
                width: 100%;
                min-height: 88px;
                display: flex;
                justify-content: center;
                align-items: center;
                margin: .6rem 0 .6rem;
            }

            .app-logo img {
                width: 220px;
                max-width: min(60vw, 220px);
                max-height: 86px;
                object-fit: contain;
            }

            .app-title {
                text-align: center;
                color: #17365D;
                font-size: 2rem;
                font-weight: 850;
                margin: .1rem 0;
            }

            .app-subtitle {
                text-align: center;
                color: #64748B;
                font-size: 1rem;
                margin-bottom: 1.2rem;
            }

            .dictionary-card {
                border: 1px solid #D0D5DD;
                border-radius: 14px;
                padding: 14px 16px;
                background: #FFFFFF;
                height: 100%;
            }

            .dictionary-title {
                color: #17365D;
                font-size: 1rem;
                font-weight: 850;
            }

            .dictionary-caption {
                color: #64748B;
                font-size: .85rem;
                margin-top: 4px;
            }

            .metric-card {
                border: 1px solid #D0D5DD;
                border-radius: 13px;
                padding: 12px 14px;
                background: #FFFFFF;
                height: 100%;
            }

            .metric-label {
                color: #64748B;
                font-size: .76rem;
                font-weight: 750;
                text-transform: uppercase;
            }

            .metric-value {
                color: #17365D;
                font-size: 1.45rem;
                font-weight: 850;
                margin-top: 3px;
            }

            .source-pill {
                display: inline-block;
                padding: 4px 9px;
                border-radius: 999px;
                background: #EFF6FF;
                color: #175CD3;
                border: 1px solid #BFDBFE;
                font-size: .78rem;
                font-weight: 750;
            }

            div[data-testid="stDataFrame"] {
                border: 1px solid #E2E8F0;
                border-radius: 12px;
                overflow: hidden;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# LOGO Y ENCABEZADO
# ============================================================

def buscar_logo() -> Path | None:
    return next(
        (
            path
            for path in LOGO_CANDIDATES
            if path.exists() and path.is_file()
        ),
        None,
    )


def mostrar_logo() -> None:
    path = buscar_logo()

    if path is None:
        return

    try:
        if path.suffix.lower() == ".svg":
            raw = path.read_text(encoding="utf-8").encode("utf-8")
            mime = "image/svg+xml"
        else:
            raw = path.read_bytes()
            mime = (
                "image/png"
                if path.suffix.lower() == ".png"
                else "image/jpeg"
            )

        encoded = base64.b64encode(raw).decode("utf-8")

        st.markdown(
            compact_html(
                f"""
                <div class="app-logo">
                    <img src="data:{mime};base64,{encoded}" alt="Logo ENAEX">
                </div>
                """
            ),
            unsafe_allow_html=True,
        )

    except (OSError, UnicodeError) as error:
        st.warning(f"No fue posible leer el logo: {error}")


def render_header() -> None:
    mostrar_logo()

    st.markdown(
        '<div class="app-title">04 Diccionarios</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div class="app-subtitle">
            Consulta, filtra y descarga los diccionarios activos de CECO–Plantas
            y Usuarios–Cargos.
        </div>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# NORMALIZACIÓN
# ============================================================

def clean_text(value: Any) -> str:
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    text = str(value).strip()

    if text.lower() in {
        "",
        "nan",
        "none",
        "null",
        "<na>",
        "n/a",
        "no usar",
        "—",
        "-",
    }:
        return ""

    return text


def normalize_sheet_name(value: str) -> str:
    normalized = clean_text(value).upper()
    replacements = str.maketrans(
        {
            "Á": "A",
            "É": "E",
            "Í": "I",
            "Ó": "O",
            "Ú": "U",
            "Ü": "U",
            "Ñ": "N",
        }
    )
    normalized = normalized.translate(replacements)
    return re.sub(r"[^A-Z0-9]+", "_", normalized).strip("_")


def find_sheet_name(
    sheet_names: list[str],
    aliases: list[str],
) -> str | None:
    normalized_sheets = {
        normalize_sheet_name(sheet): sheet
        for sheet in sheet_names
    }

    for alias in aliases:
        normalized_alias = normalize_sheet_name(alias)

        if normalized_alias in normalized_sheets:
            return normalized_sheets[normalized_alias]

    return None


def normalize_dictionary(
    dataframe: pd.DataFrame | None,
    expected_columns: list[str],
) -> pd.DataFrame:
    if dataframe is None:
        return pd.DataFrame(columns=expected_columns)

    result = dataframe.copy()
    result.columns = [
        clean_text(column) or f"Columna_{index + 1}"
        for index, column in enumerate(result.columns)
    ]

    for column in expected_columns:
        if column not in result.columns:
            result[column] = ""

    result = result.loc[:, expected_columns].copy()

    for column in expected_columns:
        result[column] = result[column].map(
            lambda value: clean_text(value)
            if column not in {"Orden", "Desde", "Hasta"}
            else value
        )

    if "Orden" in result.columns:
        result["Orden"] = pd.to_numeric(
            result["Orden"],
            errors="coerce",
        )

    for column in ["Desde", "Hasta"]:
        if column in result.columns:
            result[column] = pd.to_numeric(
                result[column],
                errors="coerce",
            )

    result = result.dropna(axis=0, how="all")
    result = result[
        result.apply(
            lambda row: any(clean_text(value) for value in row),
            axis=1,
        )
    ]

    if "CECO" in result.columns:
        result = result[result["CECO"].map(clean_text).ne("")]
        result = result.drop_duplicates("CECO", keep="last")

    if "Correo" in result.columns:
        result = result[result["Correo"].map(clean_text).ne("")]
        result = result.drop_duplicates("Correo", keep="last")

    if "Orden" in result.columns:
        result = result.sort_values(
            ["Orden", "Desde", "Hasta"],
            na_position="last",
            kind="stable",
        )

    return result.reset_index(drop=True)


# ============================================================
# OBTENCIÓN DE DATOS
# ============================================================

def get_active_dictionaries() -> tuple[
    dict[str, pd.DataFrame],
    dict[str, str],
]:
    data = st.session_state.get(SESSION_DATA_KEY)
    file_names = st.session_state.get(SESSION_FILE_KEY, {})

    if not isinstance(data, dict):
        return {}, {}

    dictionaries: dict[str, pd.DataFrame] = {}
    sources: dict[str, str] = {}

    for key, config in DICTIONARY_CONFIG.items():
        raw = data.get(config["session_key"])
        normalized = normalize_dictionary(
            raw if isinstance(raw, pd.DataFrame) else None,
            config["columns"],
        )
        dictionaries[key] = normalized

        if isinstance(file_names, dict):
            loaded_name = clean_text(
                file_names.get(config["file_role"], "")
            )
        else:
            loaded_name = ""

        sources[key] = (
            loaded_name
            if loaded_name
            else "Archivo cargado en la versión activa"
        )

    return dictionaries, sources


# ============================================================
# BÚSQUEDA Y FILTROS
# ============================================================

def apply_global_search(
    dataframe: pd.DataFrame,
    search_text: str,
) -> pd.DataFrame:
    search = clean_text(search_text).casefold()

    if not search:
        return dataframe.copy()

    text_df = dataframe.fillna("").astype(str)

    mask = text_df.apply(
        lambda row: row.str.casefold().str.contains(
            re.escape(search),
            regex=True,
            na=False,
        ).any(),
        axis=1,
    )

    return dataframe.loc[mask].copy()


def suitable_filter_columns(
    dataframe: pd.DataFrame,
) -> list[str]:
    result: list[str] = []

    for column in dataframe.columns:
        unique_count = dataframe[column].nunique(dropna=True)

        if 1 < unique_count <= 150:
            result.append(column)

    return result


def apply_column_filters(
    dataframe: pd.DataFrame,
    dictionary_key: str,
) -> pd.DataFrame:
    result = dataframe.copy()
    columns = suitable_filter_columns(result)

    if not columns:
        st.caption(
            "No se detectaron columnas apropiadas para filtros rápidos."
        )
        return result

    selected_columns = st.multiselect(
        "Columnas para filtrar",
        options=columns,
        default=[],
        key=f"dict_{dictionary_key}_filter_columns_v02",
    )

    for column in selected_columns:
        values = sorted(
            {
                clean_text(value)
                for value in result[column].tolist()
                if clean_text(value)
            },
            key=str.casefold,
        )

        selected_values = st.multiselect(
            f"Valores de {column}",
            options=values,
            default=[],
            key=f"dict_{dictionary_key}_{column}_values_v02",
        )

        if selected_values:
            result = result[
                result[column]
                .map(clean_text)
                .isin(selected_values)
            ].copy()

    return result


# ============================================================
# EXPORTACIÓN
# ============================================================

def safe_file_part(value: str) -> str:
    return re.sub(
        r"[^A-Za-z0-9_-]+",
        "_",
        clean_text(value),
    ).strip("_") or "DICCIONARIO"


def dataframe_to_csv_bytes(dataframe: pd.DataFrame) -> bytes:
    return dataframe.to_csv(
        index=False,
        sep=";",
        lineterminator="\n",
    ).encode("utf-8-sig")


def dataframe_to_excel_bytes(
    dataframe: pd.DataFrame,
    sheet_name: str,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]

    headers = list(dataframe.columns)

    for column_index, header in enumerate(headers, start=1):
        sheet.cell(1, column_index, header)

    for row_index, row in enumerate(
        dataframe.itertuples(index=False, name=None),
        start=2,
    ):
        for column_index, value in enumerate(row, start=1):
            try:
                if pd.isna(value):
                    value = None
            except (TypeError, ValueError):
                pass

            sheet.cell(row_index, column_index, value)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="17365D",
    )
    header_font = Font(
        name="Calibri",
        size=11,
        bold=True,
        color="FFFFFF",
    )
    thin = Side(style="thin", color="D0D5DD")
    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    sheet.row_dimensions[1].height = 28
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    header_index = {
        clean_text(cell.value): cell.column
        for cell in sheet[1]
    }

    for column_name in ["Orden", "Desde", "Hasta"]:
        column_index = header_index.get(column_name)

        if column_index:
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(
                    row_index,
                    column_index,
                ).number_format = "#,##0"

    for column_index in range(1, sheet.max_column + 1):
        values = [
            clean_text(
                sheet.cell(row_index, column_index).value
            )
            for row_index in range(
                1,
                min(sheet.max_row, 250) + 1,
            )
        ]
        width = min(
            max(
                max((len(value) for value in values), default=8) + 2,
                12,
            ),
            42,
        )
        sheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width

    if sheet.max_row >= 2 and sheet.max_column >= 1:
        table_name = (
            "Tabla"
            + safe_file_part(sheet_name).replace("-", "_")
        )[:250]
        reference = (
            f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        )
        table = Table(
            displayName=table_name,
            ref=reference,
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


# ============================================================
# COMPONENTES DE INTERFAZ
# ============================================================

def escape_html(value: Any) -> str:
    return html.escape(clean_text(value))


def render_metrics(
    original: pd.DataFrame,
    filtered: pd.DataFrame,
) -> None:
    duplicate_rows = int(filtered.duplicated().sum())
    empty_cells = int(
        filtered.apply(
            lambda column: column.map(clean_text).eq("").sum()
        ).sum()
    )

    columns = st.columns(4)
    metrics = [
        ("Registros totales", len(original)),
        ("Registros visibles", len(filtered)),
        ("Columnas", len(filtered.columns)),
        ("Duplicados visibles", duplicate_rows),
    ]

    for column, (label, value) in zip(columns, metrics):
        with column:
            st.markdown(
                compact_html(
                    f"""
                    <div class="metric-card">
                        <div class="metric-label">{escape_html(label)}</div>
                        <div class="metric-value">{value:,}</div>
                    </div>
                    """
                ).replace(",", "."),
                unsafe_allow_html=True,
            )

    if empty_cells:
        st.caption(
            f"La vista contiene {empty_cells:,} celdas vacías."
            .replace(",", ".")
        )


def render_dictionary(
    dictionary_key: str,
    dataframe: pd.DataFrame,
    source_label: str,
    file_name: str,
) -> None:
    config = DICTIONARY_CONFIG[dictionary_key]

    st.markdown(
        compact_html(
            f"""
            <div class="dictionary-card">
                <div class="dictionary-title">
                    {config["icono"]} {config["titulo"]}
                </div>
                <div class="dictionary-caption">
                    {escape_html(config["description"])}
                </div>
                <div style="margin-top:9px;">
                    <span class="source-pill">
                        Fuente: {escape_html(source_label)}
                    </span>
                </div>
            </div>
            """
        ),
        unsafe_allow_html=True,
    )

    if dataframe.empty:
        st.warning(
            f"No hay registros disponibles para {config['titulo']}."
        )
        return

    search_text = st.text_input(
        "Buscar en todas las columnas",
        placeholder=(
            "Escribe un CECO, planta, correo, cargo o monto..."
        ),
        key=f"dict_{dictionary_key}_search_v02",
    )

    searched = apply_global_search(
        dataframe,
        search_text,
    )

    with st.expander(
        "Filtros avanzados",
        expanded=False,
    ):
        filtered = apply_column_filters(
            searched,
            dictionary_key,
        )

    render_metrics(dataframe, filtered)

    st.markdown("#### Resultados")

    if filtered.empty:
        st.warning(
            "No hay registros que coincidan con la búsqueda y filtros."
        )
    else:
        display = filtered.copy()

        for column in ["Orden", "Desde", "Hasta"]:
            if column in display.columns:
                display[column] = display[column].map(
                    lambda value: (
                        f"{int(value):,}".replace(",", ".")
                        if pd.notna(value)
                        else ""
                    )
                )

        st.dataframe(
            display,
            use_container_width=True,
            hide_index=True,
            height=min(
                680,
                max(280, 35 * (len(display) + 2)),
            ),
        )

    base_name = Path(
        file_name or "BBDD_LIBERACION.xlsx"
    ).stem
    timestamp = datetime.now(CHILE_TZ).strftime(
        "%Y-%m-%d_%H-%M-%S"
    )
    safe_title = safe_file_part(config["titulo"])

    excel_column, csv_column = st.columns(2)

    with excel_column:
        st.download_button(
            "⬇️ Descargar vista en Excel",
            data=dataframe_to_excel_bytes(
                filtered,
                config["titulo"],
            ),
            file_name=(
                f"{base_name}_{safe_title}_{timestamp}.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
            key=f"dict_{dictionary_key}_excel_v02",
        )

    with csv_column:
        st.download_button(
            "⬇️ Descargar vista en CSV",
            data=dataframe_to_csv_bytes(filtered),
            file_name=(
                f"{base_name}_{safe_title}_{timestamp}.csv"
            ),
            mime="text/csv",
            use_container_width=True,
            key=f"dict_{dictionary_key}_csv_v02",
        )

    with st.expander(
        "Calidad y estructura de columnas",
        expanded=False,
    ):
        info = pd.DataFrame(
            {
                "Columna": dataframe.columns,
                "Tipo detectado": [
                    str(dataframe[column].dtype)
                    for column in dataframe.columns
                ],
                "Valores no vacíos": [
                    int(
                        dataframe[column]
                        .map(clean_text)
                        .ne("")
                        .sum()
                    )
                    for column in dataframe.columns
                ],
                "Valores únicos": [
                    int(
                        dataframe[column]
                        .map(clean_text)
                        .replace("", pd.NA)
                        .nunique(dropna=True)
                    )
                    for column in dataframe.columns
                ],
            }
        )

        st.dataframe(
            info,
            use_container_width=True,
            hide_index=True,
        )


def render_no_file() -> None:
    st.warning(
        "No hay una versión activa. Primero carga los siete archivos desde "
        "**01 Cargar Versión**."
    )

    try:
        if st.button(
            "📤 Ir a 01 Cargar Versión",
            type="primary",
            use_container_width=True,
        ):
            st.switch_page("01_CARGAR_ARCHIVO_FLUJO.py")
    except Exception:
        st.info(
            "Selecciona **01 Cargar Versión** desde el menú lateral."
        )



# ============================================================
# MAIN
# ============================================================

def main() -> None:
    aplicar_estilos()
    render_header()

    data = st.session_state.get(SESSION_DATA_KEY)

    if not isinstance(data, dict):
        render_no_file()
        return

    dictionaries, sources = get_active_dictionaries()

    required_keys = set(DICTIONARY_CONFIG)
    if not required_keys.issubset(dictionaries):
        render_no_file()
        return

    cecos = dictionaries["cecos"]
    users = dictionaries["usuarios"]

    if cecos.empty and users.empty:
        st.error(
            "La versión activa no contiene los diccionarios CECO-Plantas "
            "ni Usuarios-Cargos."
        )
        return

    st.success(
        (
            f"Diccionarios activos · "
            f"**{len(cecos):,} CECO** · "
            f"**{len(users):,} usuarios**"
        ).replace(",", ".")
    )

    summary_columns = st.columns(2)

    for column, key in zip(summary_columns, ["cecos", "usuarios"]):
        config = DICTIONARY_CONFIG[key]
        dataframe = dictionaries[key]

        with column:
            st.markdown(
                compact_html(
                    f"""
                    <div class="dictionary-card">
                        <div class="dictionary-title">
                            {config["icono"]} {config["titulo"]}
                        </div>
                        <div class="dictionary-caption">
                            {len(dataframe):,} registros ·
                            {len(dataframe.columns):,} columnas
                        </div>
                        <div class="dictionary-caption">
                            {escape_html(sources[key])}
                        </div>
                    </div>
                    """
                ).replace(",", "."),
                unsafe_allow_html=True,
            )

    tab_cecos, tab_users, tab_quality = st.tabs(
        [
            "🏭 CECO Y PLANTAS",
            "👥 USUARIOS Y CARGOS",
            "✅ VALIDACIÓN",
        ]
    )

    with tab_cecos:
        render_dictionary(
            "cecos",
            cecos,
            sources["cecos"],
            sources["cecos"],
        )

    with tab_users:
        render_dictionary(
            "usuarios",
            users,
            sources["usuarios"],
            sources["usuarios"],
        )

    with tab_quality:
        st.subheader("Validación cruzada")

        flow = data.get("flujo", pd.DataFrame())
        if not isinstance(flow, pd.DataFrame) or flow.empty:
            st.info("No hay flujo consolidado para realizar validaciones.")
            return

        known_cecos = set(cecos["CECO"].map(clean_text))
        flow_cecos = set(flow["CECO"].map(clean_text))
        missing_cecos = sorted(flow_cecos - known_cecos)

        known_users = {
            clean_text(value).casefold()
            for value in users["Correo"].tolist()
            if clean_text(value)
        }
        lib_columns = [
            column for column in ["Lib1", "Lib2", "Lib3", "Lib4", "Lib5"]
            if column in flow.columns
        ]
        flow_users = {
            clean_text(value).casefold()
            for column in lib_columns
            for value in flow[column].tolist()
            if clean_text(value)
            and clean_text(value) != "Liberador Servicios"
        }
        missing_users = sorted(flow_users - known_users)

        metrics = st.columns(4)
        metrics[0].metric("CECO en flujo", len(flow_cecos))
        metrics[1].metric("CECO sin diccionario", len(missing_cecos))
        metrics[2].metric("Usuarios en flujo", len(flow_users))
        metrics[3].metric("Usuarios sin diccionario", len(missing_users))

        if missing_cecos:
            st.error("Hay CECO del flujo que no existen en el diccionario.")
            st.dataframe(
                pd.DataFrame({"CECO": missing_cecos}),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.success("Todos los CECO del flujo existen en el diccionario.")

        if missing_users:
            st.error(
                "Hay liberadores del flujo que no existen en el diccionario "
                "de usuarios."
            )
            st.dataframe(
                pd.DataFrame({"Correo": missing_users}),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.success(
                "Todos los usuarios del flujo existen en el diccionario."
            )



if __name__ == "__main__":
    main()
