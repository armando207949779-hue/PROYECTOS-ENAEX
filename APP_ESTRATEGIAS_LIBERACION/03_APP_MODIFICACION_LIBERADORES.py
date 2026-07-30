# ============================================================
# 03_APP_MODIFICACION_LIBERADORES
# APP_ESTRATEGIAS_LIBERACION
#
# Asistente de modificación basado en escenarios:
# situación → alcance → CECO/tipo/rango → pieza/acción
# → validación → vista previa → impacto global opcional
# → auditoría → Excel profesional.
#
# Fuente vigente: cinco liberadores y dos diccionarios.
# La pantalla actualiza reglas explícitas y reglas especiales,
# incluidas las filas sin CECO o con CostCenter='*'.
# ============================================================

from __future__ import annotations

import base64
import hashlib
import importlib.util
import re
import zipfile
from datetime import datetime
from zoneinfo import ZoneInfo
from html import escape
from io import BytesIO
from pathlib import Path
from textwrap import dedent
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ============================================================
# CONFIGURACIÓN
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent

LOGO_CANDIDATES = [
    PROJECT_DIR / "assets" / "logo.svg",
    BASE_DIR / "assets" / "logo.svg",
    PROJECT_DIR / "assets" / "logo.png",
    BASE_DIR / "assets" / "logo.png",
    PROJECT_DIR / "assets" / "logo.jpg",
    BASE_DIR / "assets" / "logo.jpg",
    PROJECT_DIR / "assets" / "logo.jpeg",
    BASE_DIR / "assets" / "logo.jpeg",
]

SESSION_DATA_KEY = "flujo_liberacion_data"
SESSION_FILE_KEY = "flujo_liberacion_file_name"
SESSION_FILE_BYTES_KEY = "flujo_liberacion_file_bytes"
SESSION_SOURCE_FILES_KEY = "flujo_liberacion_source_files_v05"

SESSION_WORKING_KEY = "mod_liberadores_working_df_v04"
SESSION_BACKUP_KEY = "mod_liberadores_backup_df_v04"
SESSION_SIGNATURE_KEY = "mod_liberadores_signature_v04"
SESSION_DRAFT_KEY = "mod_liberadores_draft_v04"
SESSION_HISTORY_KEY = "mod_liberadores_history_v04"
SESSION_DOWNLOAD_KEY = "mod_liberadores_download_v04"
SESSION_DOWNLOAD_NAME_KEY = "mod_liberadores_download_name_v04"
SESSION_DOWNLOAD_PARQUET_KEY = "mod_liberadores_download_parquet_v02"
SESSION_DOWNLOAD_CSV_KEY = "mod_liberadores_download_csv_v02"
SESSION_DOWNLOAD_PARQUET_NAME_KEY = "mod_liberadores_download_parquet_name_v02"
SESSION_DOWNLOAD_CSV_NAME_KEY = "mod_liberadores_download_csv_name_v02"
SESSION_DOWNLOAD_EXCEL_KEY = "mod_liberadores_download_excel_v01"
SESSION_DOWNLOAD_EXCEL_NAME_KEY = "mod_liberadores_download_excel_name_v01"

LIB_COLS = ["Lib1", "Lib2", "Lib3", "Lib4", "Lib5"]
FLOW_COLUMNS = [
    "CECO", "Planta", "Desde", "Hasta", "TipoDoc",
    "Lib1", "Lib2", "Lib3", "Lib4", "Lib5",
]

DOC_LABEL = {
    "AZNB": "Material (AZNB)",
    "AZSR": "Servicio (AZSR)",
    "AMBOS": "Ambos (AZNB + AZSR)",
}

ACTION_LABEL = {
    "mover": "Mover a otra posición",
    "reemplazar": "Reemplazar esta pieza",
    "eliminar": "Eliminar esta pieza",
    "agregar": "Agregar una pieza nueva",
}

SCENARIO_LABEL = {
    "reemplazo_ceco": "🔁 Reemplazar usuario en un CECO completo",
    "ajuste": "🛠️ Ajustar un flujo específico",
    "salida": "🚪 Reemplazar a alguien que salió de la empresa",
    "temporal": "🗓️ Cubrir una ausencia temporal",
    "orden": "↕️ Reordenar aprobadores de un tramo",
    "dotacion": "➕➖ Agregar o retirar un liberador",
}

SCENARIO_HELP = {
    "reemplazo_ceco": (
        "Reemplaza un usuario en todas sus apariciones del CECO seleccionado, "
        "sin depender del rango. Puede incluir el CECO gemelo EMTS y aplicar "
        "el cambio a Material, Servicio o ambos tipos."
    ),
    "ajuste": (
        "Permite mover, reemplazar, agregar o eliminar una pieza "
        "en un CECO, tipo y rango específicos."
    ),
    "salida": (
        "Busca todas las apariciones de una persona y la reemplaza "
        "en todos los CECO de la base."
    ),
    "temporal": (
        "Reemplaza una persona solo en el tramo seleccionado, "
        "sin afectar sus demás participaciones."
    ),
    "orden": (
        "Cambia el orden de aprobación dentro de un tramo, "
        "sin modificar las personas."
    ),
    "dotacion": (
        "Agrega un nuevo liberador o retira uno existente "
        "en un tramo específico."
    ),
}

LS_LABEL = "Liberador Servicios"
LS_GROUP = "Liberacion Servicios"
LEVELS = (1, 2, 3, 4, 5)
BASE_RULE_COLUMNS = [
    "CompanyCode", "BillingAddress", "AccountCategory", "CostCenter",
    "cus_POClasedeDocumento", "PurchaseGroup",
    "TotalCost Bajo", "TotalCost Alto",
]
EXPORT_LEVEL_COLUMNS = [
    *BASE_RULE_COLUMNS, "Group", "User", "Required", "Tooltip",
]

# Zona horaria oficial de Santiago de Chile. ZoneInfo aplica
# automáticamente los cambios de horario de verano/invierno.
CHILE_TZ = ZoneInfo("America/Santiago")


# ============================================================
# ESTILOS
# ============================================================

def compact_html(value: str) -> str:
    return re.sub(r">\s+<", "><", dedent(value).strip())


def aplicar_estilos() -> None:
    st.markdown(
        """
        <style>
            .stMainBlockContainer,
            .block-container {
                padding-top: 6.5rem !important;
                padding-bottom: 3rem !important;
            }

            .app-logo {
                width: 100%;
                min-height: 88px;
                display: flex;
                justify-content: center;
                align-items: center;
                margin: .6rem 0 .6rem;
                overflow: visible;
            }

            .app-logo img {
                width: 220px;
                max-width: min(60vw, 220px);
                max-height: 86px;
                object-fit: contain;
                display: block;
            }

            .app-title {
                text-align: center;
                color: #17365D;
                font-size: 2rem;
                font-weight: 850;
                margin: .1rem 0;
            }

            .app-subtitle {
                text-align: center;
                color: #64748B;
                font-size: 1rem;
                margin-bottom: 1.2rem;
            }

            .question-card {
                border: 2px solid #93C5FD;
                border-radius: 16px;
                padding: 16px 18px;
                background: linear-gradient(135deg, #EFF6FF 0%, #F8FAFC 100%);
                margin: .8rem 0 1rem;
            }

            .question-number {
                display: inline-flex;
                width: 30px;
                height: 30px;
                align-items: center;
                justify-content: center;
                border-radius: 999px;
                background: #17365D;
                color: #FFFFFF;
                font-weight: 850;
                margin-right: 8px;
            }

            .question-title {
                color: #17365D;
                font-size: 1.15rem;
                font-weight: 850;
            }

            .question-help {
                color: #64748B;
                font-size: .9rem;
                margin-top: 6px;
            }

            .flow-card {
                min-width: 175px;
                max-width: 230px;
                border-radius: 13px;
                padding: 12px;
                font-family: Arial, sans-serif;
            }

            .comparison-grid {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(310px, 1fr));
                gap: 14px;
                margin-top: 10px;
            }

            .comparison-before {
                padding: 14px;
                border-radius: 14px;
                border: 1px solid #FECACA;
                background: #FEF2F2;
            }

            .comparison-after {
                padding: 14px;
                border-radius: 14px;
                border: 1px solid #BBF7D0;
                background: #F0FDF4;
            }

            .metric-card {
                border: 1px solid #D0D5DD;
                border-radius: 13px;
                padding: 13px 15px;
                background: #FFFFFF;
                height: 100%;
            }

            .metric-label {
                color: #64748B;
                font-size: .78rem;
                font-weight: 750;
                text-transform: uppercase;
            }

            .metric-value {
                color: #17365D;
                font-size: 1.5rem;
                font-weight: 850;
                margin-top: 4px;
            }

            .replacement-note {
                border: 1px solid #F59E0B;
                border-left: 5px solid #F59E0B;
                background: #FFFBEB;
                color: #92400E;
                border-radius: 12px;
                padding: 12px 14px;
                margin: 8px 0 12px;
                font-weight: 650;
            }

            div[data-testid="stDataFrame"] {
                border: 1px solid #E2E8F0;
                border-radius: 12px;
                overflow: hidden;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def buscar_logo() -> Path | None:
    return next(
        (path for path in LOGO_CANDIDATES if path.exists() and path.is_file()),
        None,
    )


def mostrar_logo() -> None:
    path = buscar_logo()
    if path is None:
        return

    try:
        if path.suffix.lower() == ".svg":
            raw = path.read_text(encoding="utf-8").encode("utf-8")
            mime = "image/svg+xml"
        else:
            raw = path.read_bytes()
            mime = "image/png" if path.suffix.lower() == ".png" else "image/jpeg"

        encoded = base64.b64encode(raw).decode("utf-8")
        st.markdown(
            compact_html(
                f"""
                <div class="app-logo">
                    <img src="data:{mime};base64,{encoded}" alt="Logo ENAEX">
                </div>
                """
            ),
            unsafe_allow_html=True,
        )
    except (OSError, UnicodeError):
        st.warning(f"No fue posible leer el logo: {path.name}")


def render_header() -> None:
    mostrar_logo()
    st.markdown(
        '<div class="app-title">03 Modificación de Liberadores</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div class="app-subtitle">
            Modifica reglas con y sin CECO y descarga cinco liberadores, dos diccionarios y cambios.
        </div>
        """,
        unsafe_allow_html=True,
    )


def question(number: int, title: str, help_text: str) -> None:
    st.markdown(
        compact_html(
            f"""
            <div class="question-card">
                <div>
                    <span class="question-number">{number}</span>
                    <span class="question-title">{escape(title)}</span>
                </div>
                <div class="question-help">{escape(help_text)}</div>
            </div>
            """
        ),
        unsafe_allow_html=True,
    )


# ============================================================
# NORMALIZACIÓN
# ============================================================

def validate_flow_schema(data: dict[str, Any]) -> pd.DataFrame:
    """Valida la versión activa reconstruida desde cinco archivos."""
    if not isinstance(data, dict):
        raise ValueError(
            "No existe una versión activa. Carga los cinco archivos "
            "desde 01 Cargar Liberadores."
        )

    flow = data.get("flujo")
    if not isinstance(flow, pd.DataFrame) or flow.empty:
        raise ValueError("La versión activa no contiene un flujo válido.")

    missing = [column for column in FLOW_COLUMNS if column not in flow.columns]
    if missing:
        raise ValueError(
            "El flujo reconstruido no tiene la estructura requerida. "
            f"Faltan: {', '.join(missing)}."
        )

    liberadores = data.get("liberadores")
    if not isinstance(liberadores, dict):
        raise ValueError(
            "La versión activa no conserva los cinco archivos de liberadores."
        )

    missing_levels = [
        level for level in LEVELS
        if not isinstance(liberadores.get(level), pd.DataFrame)
    ]
    if missing_levels:
        raise ValueError(
            "Faltan niveles: "
            + ", ".join(f"Liberador {level}" for level in missing_levels)
            + "."
        )

    return flow



def clean_text(value: Any) -> str:
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none", "null", "<na>", "n/a", "no usar", "—", "-"} else text


def strip_user(value: Any) -> str:
    text = clean_text(value)
    if not text or text == LS_LABEL:
        return text

    match = re.match(r"^(.*?)(?:\s+\([^)]+\))?$", text)
    return match.group(1).strip() if match else text


def email_key(value: Any) -> str:
    return strip_user(value).lower()


def parse_bound(value: Any, low: bool = True) -> float:
    text = clean_text(value)
    if not text or text == "*":
        return 0.0 if low else 1e18

    normalized = text.replace(" ", "")

    if "." in normalized and "," in normalized:
        if normalized.rfind(",") > normalized.rfind("."):
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", "")
    elif normalized.count(".") > 1:
        normalized = normalized.replace(".", "")
    elif normalized.count(",") > 1:
        normalized = normalized.replace(",", "")
    elif "," in normalized:
        left, right = normalized.rsplit(",", 1)
        normalized = left + right if len(right) == 3 else left + "." + right
    elif "." in normalized:
        left, right = normalized.rsplit(".", 1)
        if len(right) == 3 and left.replace("-", "").isdigit():
            normalized = left + right

    try:
        return float(normalized)
    except ValueError:
        return 0.0 if low else 1e18


def fmt_bound(value: Any) -> str:
    number = parse_bound(value, low=False)
    if number >= 1e12:
        return "1E+12"
    return f"{int(number):,}".replace(",", ".")


def normalize_flow(df: pd.DataFrame) -> pd.DataFrame:
    flow = df.copy()
    flow.columns = [str(column).strip() for column in flow.columns]

    for column in FLOW_COLUMNS:
        if column not in flow.columns:
            flow[column] = ""

    flow = flow.loc[:, FLOW_COLUMNS].copy()

    for column in ["CECO", "Planta"]:
        flow[column] = flow[column].map(clean_text)

    flow["TipoDoc"] = flow["TipoDoc"].map(clean_text).str.upper()

    for column in LIB_COLS:
        flow[column] = flow[column].map(strip_user)


    flow = flow[
        flow["CECO"].ne("")
        & flow["TipoDoc"].isin(["AZNB", "AZSR"])
    ].reset_index(drop=True)

    flow.insert(0, "_ID_FILA", range(1, len(flow) + 1))
    return flow


def libs_from_row(row: pd.Series) -> list[str]:
    return [
        strip_user(row.get(column, ""))
        for column in LIB_COLS
        if strip_user(row.get(column, ""))
    ]


def libs_padded(values: list[str]) -> list[str]:
    cleaned = [
        strip_user(value)
        for value in values
        if strip_user(value)
    ]
    return (cleaned + [""] * 5)[:5]


def cargo_dictionary(data: dict[str, pd.DataFrame]) -> dict[str, str]:
    users = data.get("dic_users", pd.DataFrame())
    if not isinstance(users, pd.DataFrame) or users.empty:
        return {}

    email_col = next(
        (
            column
            for column in users.columns
            if str(column).strip().lower() in {"correo", "email", "mail"}
        ),
        None,
    )
    cargo_col = next(
        (
            column
            for column in users.columns
            if str(column).strip().lower() in {"cargo", "rol", "role"}
        ),
        None,
    )

    if email_col is None:
        return {}

    result: dict[str, str] = {}
    for _, row in users.iterrows():
        email = strip_user(row.get(email_col, ""))
        cargo = clean_text(row.get(cargo_col, "")) if cargo_col else ""
        if email:
            result[email.lower()] = cargo
    return result


def display_user(value: Any, data: dict[str, pd.DataFrame]) -> str:
    email = strip_user(value)
    if not email:
        return ""
    if email == LS_LABEL:
        return LS_LABEL

    cargo = cargo_dictionary(data).get(email.lower(), "")
    return f"{email} ({cargo})" if cargo else email


def unique_users(flow: pd.DataFrame) -> list[str]:
    return sorted(
        {
            strip_user(value)
            for column in LIB_COLS
            for value in flow[column].tolist()
            if strip_user(value) and strip_user(value) != LS_LABEL
        },
        key=str.lower,
    )



def normalize_plant_family(value: Any) -> str:
    """Normaliza la planta para identificar su CECO base y su gemelo EMTS."""
    text = clean_text(value).upper()
    text = re.sub(r"\bEMTS\b", "", text)
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def find_twin_cecos(
    flow: pd.DataFrame,
    selected_ceco: str,
) -> list[dict[str, str]]:
    """Busca CECO distintos cuya planta pertenece a la misma familia."""
    selected_rows = flow[flow["CECO"].eq(selected_ceco)]
    selected_plants = [
        clean_text(value)
        for value in selected_rows["Planta"].tolist()
        if clean_text(value)
    ]
    families = {
        normalize_plant_family(value)
        for value in selected_plants
        if normalize_plant_family(value)
    }

    if not families:
        return []

    candidates = (
        flow[["CECO", "Planta"]]
        .drop_duplicates()
        .sort_values(["Planta", "CECO"], kind="stable")
    )

    twins: list[dict[str, str]] = []
    seen: set[str] = set()

    for _, row in candidates.iterrows():
        ceco = clean_text(row.get("CECO"))
        planta = clean_text(row.get("Planta"))

        if not ceco or ceco == selected_ceco or ceco in seen:
            continue

        if normalize_plant_family(planta) not in families:
            continue

        seen.add(ceco)
        twins.append({"ceco": ceco, "planta": planta})

    return twins


def same_range_mask(
    dataframe: pd.DataFrame,
    desde: Any,
    hasta: Any,
) -> pd.Series:
    """Compara rangos usando valores numéricos normalizados."""
    target_from = parse_bound(desde, low=True)
    target_until = parse_bound(hasta, low=False)

    return (
        dataframe["Desde"].map(lambda value: parse_bound(value, low=True)).eq(target_from)
        & dataframe["Hasta"].map(lambda value: parse_bound(value, low=False)).eq(target_until)
    )


def row_flow_signature(row: pd.Series) -> tuple[str, ...]:
    return tuple(libs_padded(libs_from_row(row)))


def is_valid_email(value: Any) -> bool:
    """Valida correos simples; Liberador Servicios es una excepción válida."""
    text = strip_user(value)
    if text == LS_LABEL:
        return True
    if not text:
        return False
    return bool(
        re.fullmatch(
            r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}",
            text,
        )
    )


def validate_flow_result(
    libs: list[str],
    *,
    allow_empty: bool = False,
) -> list[str]:
    """Valida el flujo borrador antes de permitir su guardado."""
    errors: list[str] = []
    cleaned = [strip_user(value) for value in libs if strip_user(value)]

    if not cleaned and not allow_empty:
        errors.append(
            "El tramo no puede quedar sin liberadores. "
            "Agrega al menos una persona antes de guardar."
        )

    if len(cleaned) > 5:
        errors.append("El flujo no puede superar 5 liberadores.")

    normalized = [email_key(value) for value in cleaned]
    if len(normalized) != len(set(normalized)):
        errors.append("El flujo contiene liberadores duplicados.")

    invalid = [
        value
        for value in cleaned
        if value != LS_LABEL and not is_valid_email(value)
    ]
    if invalid:
        errors.append(
            "Existen correos con formato no válido: "
            + ", ".join(invalid)
        )

    return errors



# ============================================================
# ESTADO
# ============================================================

def source_signature(
    file_name: Any,
    file_bytes: Any,
    rows: int,
) -> str:
    """Firma estable de la versión formada por cinco archivos."""
    name_parts: list[str] = []
    byte_parts: list[str] = []

    if isinstance(file_name, dict):
        for level in LEVELS:
            name_parts.append(
                f"{level}:{clean_text(file_name.get(level, ''))}"
            )
    else:
        name_parts.append(clean_text(file_name))

    if isinstance(file_bytes, dict):
        for level in LEVELS:
            raw = file_bytes.get(f"liberador_{level}", b"")
            if not isinstance(raw, (bytes, bytearray)):
                raw = b""
            digest = hashlib.sha1(bytes(raw)[:100000]).hexdigest()
            byte_parts.append(f"{level}:{len(raw)}:{digest}")
    elif isinstance(file_bytes, (bytes, bytearray)):
        digest = hashlib.sha1(bytes(file_bytes)[:100000]).hexdigest()
        byte_parts.append(f"{len(file_bytes)}:{digest}")
    else:
        byte_parts.append("sin-bytes")

    return "|".join([*name_parts, *byte_parts, str(rows)])



def default_draft() -> dict[str, Any]:
    return {
        "ceco": "",
        "doc": "",
        "row_id": None,
        "target_row_ids": [],
        "target_cecos": [],
        "target_docs": [],
        "libs_before": [],
        "libs_after": [],
        "replaced_from": "",
        "replaced_to": "",
        "last_message": "",
    }


def initialize_state(
    data: dict[str, pd.DataFrame],
    file_name: str,
    file_bytes: bytes,
) -> None:
    flow = normalize_flow(validate_flow_schema(data))
    signature = source_signature(file_name, file_bytes, len(flow))

    if (
        SESSION_WORKING_KEY not in st.session_state
        or st.session_state.get(SESSION_SIGNATURE_KEY) != signature
    ):
        st.session_state[SESSION_WORKING_KEY] = flow
        st.session_state[SESSION_BACKUP_KEY] = flow.copy(deep=True)
        st.session_state[SESSION_SIGNATURE_KEY] = signature
        st.session_state[SESSION_DRAFT_KEY] = default_draft()
        loaded_changes = data.get("cambios", pd.DataFrame())
        st.session_state[SESSION_HISTORY_KEY] = (
            loaded_changes.to_dict("records")
            if isinstance(loaded_changes, pd.DataFrame)
            and not loaded_changes.empty
            else []
        )
        st.session_state.pop(SESSION_DOWNLOAD_KEY, None)
        st.session_state.pop(SESSION_DOWNLOAD_NAME_KEY, None)


def get_working_flow() -> pd.DataFrame:
    value = st.session_state.get(SESSION_WORKING_KEY)
    if isinstance(value, pd.DataFrame):
        return value.copy(deep=True)
    return pd.DataFrame(columns=["_ID_FILA", *FLOW_COLUMNS])


def set_working_flow(flow: pd.DataFrame) -> None:
    result = flow.copy(deep=True).reset_index(drop=True)
    result["_ID_FILA"] = range(1, len(result) + 1)
    st.session_state[SESSION_WORKING_KEY] = result

    data = st.session_state.get(SESSION_DATA_KEY)
    if isinstance(data, dict):
        updated = dict(data)
        updated["flujo"] = (
            result.drop(columns=["_ID_FILA"], errors="ignore")
            .loc[:, FLOW_COLUMNS]
            .copy()
        )
        st.session_state[SESSION_DATA_KEY] = updated


def get_draft() -> dict[str, Any]:
    value = st.session_state.get(SESSION_DRAFT_KEY)
    if not isinstance(value, dict):
        value = default_draft()
        st.session_state[SESSION_DRAFT_KEY] = value
    return dict(value)


def set_draft(draft: dict[str, Any]) -> None:
    st.session_state[SESSION_DRAFT_KEY] = dict(draft)


def reset_draft() -> None:
    st.session_state[SESSION_DRAFT_KEY] = default_draft()


# ============================================================
# TABLAS Y FLUJOS
# ============================================================

def style_flow_table(df: pd.DataFrame):
    visible_columns = [
        "CECO", "Planta", "Desde", "Hasta", "TipoDoc",
        "Lib1", "Lib2", "Lib3", "Lib4", "Lib5",
    ]
    visible = df.loc[:, visible_columns].copy()

    def style_row(row: pd.Series) -> list[str]:
        if row["TipoDoc"] == "AZNB":
            style = "background-color:#FFF1F0;color:#7A271A;"
        elif row["TipoDoc"] == "AZSR":
            style = "background-color:#EFF8FF;color:#1849A9;"
        else:
            style = ""
        return [style] * len(row)

    return visible.style.apply(style_row, axis=1).format(
        {
            "Desde": lambda value: fmt_bound(value),
            "Hasta": lambda value: fmt_bound(value),
        }
    )


def flow_html(
    libs: list[str],
    data: dict[str, pd.DataFrame],
    title: str,
) -> str:
    if not libs:
        return (
            f"<div style='font-family:Arial,sans-serif;'>"
            f"<b>{escape(title)}</b>"
            f"<p style='color:#B42318;'>Sin liberadores.</p></div>"
        )

    parts: list[str] = []

    for index, user in enumerate(libs):
        parts.append(
            f"""
            <div class="flow-card"
                 style="background:#F8FAFC;border:2px dashed #94A3B8;">
                <div style="font-size:11px;color:#64748B;font-weight:750;">
                    Liberador {index + 1}
                </div>
                <div style="
                    font-size:12px;
                    font-weight:750;
                    color:#17365D;
                    margin-top:7px;
                    overflow-wrap:anywhere;
                ">
                    {escape(display_user(user, data))}
                </div>
            </div>
            """
        )

        if index < len(libs) - 1:
            parts.append(
                "<div style='font-size:21px;color:#94A3B8;font-weight:700;'>→</div>"
            )

    return compact_html(
        f"""
        <div style="font-family:Arial,sans-serif;">
            <div style="font-weight:850;color:#17365D;margin-bottom:9px;">
                {escape(title)}
            </div>
            <div style="display:flex;flex-wrap:wrap;align-items:center;gap:7px;">
                {''.join(parts)}
            </div>
        </div>
        """
    )


def render_comparison(
    before: list[str],
    after: list[str],
    data: dict[str, pd.DataFrame],
) -> None:
    st.markdown(
        compact_html(
            f"""
            <div class="comparison-grid">
                <div class="comparison-before">
                    {flow_html(before, data, "ANTES")}
                </div>
                <div class="comparison-after">
                    {flow_html(after, data, "DESPUÉS · BORRADOR")}
                </div>
            </div>
            """
        ),
        unsafe_allow_html=True,
    )


# ============================================================
# ACCIONES
# ============================================================

def apply_action(
    action: str,
    libs: list[str],
    selected_index: int | None,
    destination_index: int | None,
    new_value: str,
) -> tuple[list[str], str, str, str]:
    result = list(libs)
    replaced_from = ""
    replaced_to = ""

    if action == "agregar":
        if len(result) >= 5:
            raise ValueError("El flujo ya tiene el máximo de 5 liberadores.")
        if not new_value:
            raise ValueError("Indica quién deseas agregar.")
        if email_key(new_value) in [email_key(value) for value in result]:
            raise ValueError("La persona ya existe en el flujo.")

        result.append(new_value)
        return (
            result,
            f"Agregado como Liberador {len(result)}: {new_value}.",
            "",
            "",
        )

    if selected_index is None:
        raise ValueError("Selecciona una pieza.")

    if selected_index < 0 or selected_index >= len(result):
        raise ValueError("La pieza seleccionada no es válida.")

    if action == "eliminar":
        removed = result.pop(selected_index)
        return result, f"Eliminado: {removed}.", "", ""

    if action == "mover":
        if destination_index is None:
            raise ValueError("Indica a qué posición deseas mover la pieza.")
        if destination_index == selected_index:
            raise ValueError("La pieza ya está en esa posición.")

        piece = result.pop(selected_index)
        result.insert(destination_index, piece)
        return (
            result,
            f"Movido {piece} a Liberador {destination_index + 1}.",
            "",
            "",
        )

    if action == "reemplazar":
        if not new_value:
            raise ValueError("Indica quién entra.")

        old_value = result[selected_index]
        keys = [email_key(value) for value in result]

        if (
            email_key(new_value) in keys
            and email_key(new_value) != email_key(old_value)
        ):
            other_index = keys.index(email_key(new_value))
            result[selected_index], result[other_index] = (
                result[other_index],
                result[selected_index],
            )
            return (
                result,
                (
                    "La persona ya estaba en el flujo. "
                    f"Se intercambiaron Liberador {selected_index + 1} "
                    f"y Liberador {other_index + 1}."
                ),
                "",
                "",
            )

        result[selected_index] = new_value
        replaced_from = old_value
        replaced_to = new_value

        return (
            result,
            (
                f"Reemplazo en Liberador {selected_index + 1}: "
                f"{old_value} → {new_value}."
            ),
            replaced_from,
            replaced_to,
        )

    raise ValueError("Selecciona una acción válida.")


def occurrences_of_person(
    flow: pd.DataFrame,
    person: str,
) -> pd.DataFrame:
    target = email_key(person)
    records: list[dict[str, Any]] = []

    if not target:
        return pd.DataFrame()

    for _, row in flow.iterrows():
        for column in LIB_COLS:
            current = strip_user(row[column])
            if email_key(current) != target:
                continue

            records.append(
                {
                    "_ID_FILA": int(row["_ID_FILA"]),
                    "CECO": row["CECO"],
                    "Planta": row["Planta"],
                    "Desde": row["Desde"],
                    "Hasta": row["Hasta"],
                    "TipoDoc": row["TipoDoc"],
                    "Campo": column,
                    "ValorAntes": current,
                }
            )

    return pd.DataFrame(records)


# ============================================================
# EXPORTACIÓN DE LOS CINCO ARCHIVOS
# ============================================================

def download_name(original_name: Any) -> str:
    timestamp = datetime.now(CHILE_TZ).strftime("%Y-%m-%d_%H-%M-%S")
    return f"LIBERADORES_ACTUALIZADOS_{timestamp}.zip"


def sanitize_excel_value(value: Any) -> Any:
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def remove_existing_tables(sheet) -> None:
    for table_name in list(sheet.tables.keys()):
        del sheet.tables[table_name]


def professional_sheet_format(
    sheet,
    *,
    table_name: str,
    header_fill: str = "17365D",
    tab_color: str | None = None,
) -> None:
    if sheet.max_row < 1 or sheet.max_column < 1:
        return

    if tab_color:
        sheet.sheet_properties.tabColor = tab_color

    header_font = Font(
        name="Calibri", size=11, bold=True, color="FFFFFF"
    )
    header_pattern = PatternFill(
        fill_type="solid", fgColor=header_fill
    )
    thin_gray = Side(style="thin", color="D0D5DD")
    body_border = Border(
        left=thin_gray, right=thin_gray,
        top=thin_gray, bottom=thin_gray,
    )

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_pattern
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = body_border

    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(
                vertical="top", wrap_text=True
            )
            cell.border = body_border
            cell.number_format = "@"

    for column_index in range(1, sheet.max_column + 1):
        letter = get_column_letter(column_index)
        values = [
            str(sheet.cell(row_index, column_index).value or "")
            for row_index in range(1, min(sheet.max_row, 250) + 1)
        ]
        width = min(max(max(map(len, values), default=8) + 2, 11), 42)
        header = str(sheet.cell(1, column_index).value or "")

        if header in {"User", "Tooltip"}:
            width = max(width, 32)
        elif header in {
            "CostCenter", "cus_POClasedeDocumento", "PurchaseGroup"
        }:
            width = max(width, 18)

        sheet.column_dimensions[letter].width = width

    remove_existing_tables(sheet)

    if sheet.max_row >= 2:
        safe_name = re.sub(r"[^A-Za-z0-9_]", "_", table_name)
        reference = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        table = Table(displayName=safe_name[:250], ref=reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def write_dataframe_to_sheet(sheet, dataframe: pd.DataFrame) -> None:
    for row_index, values in enumerate(
        dataframe_to_rows(dataframe, index=False, header=True),
        start=1,
    ):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(
                row=row_index,
                column=column_index,
                value=sanitize_excel_value(value),
            )


def normalized_level_frame(frame: pd.DataFrame, level: int) -> pd.DataFrame:
    """Prepara un nivel sin eliminar reglas especiales ni columnas funcionales."""
    result = frame.copy()

    for column in EXPORT_LEVEL_COLUMNS:
        if column not in result.columns:
            result[column] = ""

    result = result.loc[:, EXPORT_LEVEL_COLUMNS].copy()

    for column in EXPORT_LEVEL_COLUMNS:
        result[column] = result[column].map(clean_text)

    result["Required"] = result["Required"].replace("", "TRUE")
    result["Tooltip"] = result["Tooltip"].replace(
        "", f"Liberador {level} - Directa ENAEX"
    )
    return result


def range_key(value: Any, *, low: bool) -> float:
    return parse_bound(value, low=low)


def replace_user_in_special_rules(
    data: dict[str, Any],
    old_user: str,
    new_user: str,
    actor: str,
    reason: str,
    timestamp: str,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """
    Reemplaza un usuario en reglas que no forman parte del flujo CECO.

    Incluye filas con CostCenter vacío o '*', documentos distintos de
    AZNB/AZSR y otras reglas especiales que el flujo consolidado no contiene.
    """
    updated_data = dict(data)
    original_frames = data.get("liberadores", {})

    if not isinstance(original_frames, dict):
        return updated_data, []

    old_key = email_key(old_user)
    replacement = strip_user(new_user)
    updated_frames: dict[int, pd.DataFrame] = {}
    changes: list[dict[str, Any]] = []

    for level in LEVELS:
        original = normalized_level_frame(
            original_frames.get(level, pd.DataFrame()),
            level,
        )

        special_mask = (
            original["CostCenter"].isin({"", "*"})
            | ~original["cus_POClasedeDocumento"].isin(
                {"AZNB", "AZSR"}
            )
        )

        for row_index, row in original[special_mask].iterrows():
            current = strip_user(row.get("User", ""))

            if email_key(current) != old_key:
                continue

            original.at[row_index, "User"] = replacement

            ceco_value = clean_text(row.get("CostCenter", ""))
            changes.append(
                {
                    "FechaHora": timestamp,
                    "Usuario": actor or "anonimo",
                    "CECO": ceco_value or "[SIN CECO]",
                    "Desde": clean_text(
                        row.get("TotalCost Bajo", "")
                    ),
                    "Hasta": clean_text(
                        row.get("TotalCost Alto", "")
                    ),
                    "TipoDoc": clean_text(
                        row.get(
                            "cus_POClasedeDocumento",
                            "",
                        )
                    )
                    or "*",
                    "Campo": f"Liberador {level} · User",
                    "ValorAntes": current,
                    "ValorDespues": replacement,
                    "Nota": (
                        (reason or "Reemplazo global")
                        + " | regla especial sin CECO explícito"
                    ),
                }
            )

        updated_frames[level] = original.reset_index(drop=True)

    updated_data["liberadores"] = updated_frames
    for level in LEVELS:
        updated_data[f"liberador_{level}"] = updated_frames[level]

    return updated_data, changes


def special_occurrences_of_person(
    data: dict[str, Any],
    person: str,
) -> pd.DataFrame:
    """Lista las apariciones de una persona en reglas especiales."""
    frames = data.get("liberadores", {})
    target = email_key(person)
    records: list[dict[str, Any]] = []

    if not target or not isinstance(frames, dict):
        return pd.DataFrame()

    for level in LEVELS:
        frame = normalized_level_frame(
            frames.get(level, pd.DataFrame()),
            level,
        )
        special_mask = (
            frame["CostCenter"].isin({"", "*"})
            | ~frame["cus_POClasedeDocumento"].isin(
                {"AZNB", "AZSR"}
            )
        )

        for _, row in frame[special_mask].iterrows():
            current = strip_user(row.get("User", ""))
            if email_key(current) != target:
                continue

            records.append(
                {
                    "Nivel": f"Liberador {level}",
                    "CECO": (
                        clean_text(row.get("CostCenter", ""))
                        or "[SIN CECO]"
                    ),
                    "CompanyCode": clean_text(
                        row.get("CompanyCode", "")
                    ),
                    "TipoDoc": (
                        clean_text(
                            row.get(
                                "cus_POClasedeDocumento",
                                "",
                            )
                        )
                        or "*"
                    ),
                    "PurchaseGroup": clean_text(
                        row.get("PurchaseGroup", "")
                    ),
                    "Desde": clean_text(
                        row.get("TotalCost Bajo", "")
                    ),
                    "Hasta": clean_text(
                        row.get("TotalCost Alto", "")
                    ),
                    "Persona actual": current,
                }
            )

    return pd.DataFrame(records)


def flow_to_level_frames(
    flow: pd.DataFrame,
    original_frames: dict[int, pd.DataFrame],
) -> dict[int, pd.DataFrame]:
    """
    Proyecta el flujo editado hacia los cinco archivos.

    Las reglas especiales (CostCenter='*', AccountCategory='V',
    PurchaseGroup específico, etc.) se conservan intactas. Solo se reemplazan
    las reglas explícitas por CECO/tipo/rango.
    """
    clean_flow = (
        flow.drop(columns=["_ID_FILA"], errors="ignore")
        .loc[:, FLOW_COLUMNS]
        .copy()
    )

    outputs: dict[int, pd.DataFrame] = {}

    for level in LEVELS:
        original = normalized_level_frame(
            original_frames.get(level, pd.DataFrame()),
            level,
        )

        special_mask = (
            original["CostCenter"].isin({"", "*"})
            | ~original["cus_POClasedeDocumento"].isin({"AZNB", "AZSR"})
        )
        special = original[special_mask].copy()

        generated_rows: list[dict[str, Any]] = []

        for _, row in clean_flow.iterrows():
            liberator = strip_user(row.get(f"Lib{level}", ""))
            if not liberator:
                continue

            ceco = clean_text(row.get("CECO", ""))
            doc = clean_text(row.get("TipoDoc", "")).upper()

            if not ceco or doc not in {"AZNB", "AZSR"}:
                continue

            low = parse_bound(row.get("Desde"), low=True)
            high = parse_bound(row.get("Hasta"), low=False)

            generated = {
                "CompanyCode": ceco[:4],
                "BillingAddress": "*",
                "AccountCategory": "*",
                "CostCenter": ceco,
                "cus_POClasedeDocumento": doc,
                "PurchaseGroup": "*",
                "TotalCost Bajo": (
                    "1" if low <= 1 else str(int(low))
                ),
                "TotalCost Alto": (
                    "*" if high >= 1e12 else str(int(high))
                ),
                "Group": "",
                "User": "",
                "Required": "TRUE",
                "Tooltip": f"Liberador {level} - Directa ENAEX",
            }

            if liberator == LS_LABEL:
                generated["Group"] = LS_GROUP
            else:
                generated["User"] = liberator

            generated_rows.append(generated)

        generated_df = pd.DataFrame(
            generated_rows,
            columns=EXPORT_LEVEL_COLUMNS,
        )

        # Las reglas especiales se conservan exactamente, incluso cuando el
        # archivo histórico contiene duplicados deliberados.
        generated_df = generated_df.drop_duplicates(
            subset=[*BASE_RULE_COLUMNS, "Group", "User"],
            keep="first",
        )

        result = pd.concat(
            [special, generated_df],
            ignore_index=True,
        ).reset_index(drop=True)

        outputs[level] = result

    return outputs


def build_level_excel(
    level: int,
    dataframe: pd.DataFrame,
    history: list[dict[str, Any]],
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    write_dataframe_to_sheet(sheet, dataframe)
    professional_sheet_format(
        sheet,
        table_name=f"TablaLiberador{level}",
        header_fill="17365D",
        tab_color="175CD3",
    )

    if history:
        history_sheet = workbook.create_sheet("Cambios")
        history_df = pd.DataFrame(history)
        write_dataframe_to_sheet(history_sheet, history_df)
        professional_sheet_format(
            history_sheet,
            table_name=f"TablaCambiosL{level}",
            header_fill="7F1D1D",
            tab_color="B42318",
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def available_parquet_engine() -> str | None:
    if importlib.util.find_spec("pyarrow") is not None:
        return "pyarrow"
    if importlib.util.find_spec("fastparquet") is not None:
        return "fastparquet"
    return None


def chile_modification_timestamp() -> str:
    """Fecha y hora de modificación en Santiago, Chile."""
    return datetime.now(CHILE_TZ).strftime("%Y-%m-%d_%H-%M-%S")


def clean_export_frame(dataframe: pd.DataFrame) -> pd.DataFrame:
    return dataframe.drop(
        columns=[
            "_DesdeNum",
            "_HastaNum",
            "_Nivel",
            "_Liberador",
            "_ID_FILA",
        ],
        errors="ignore",
    ).copy()


def synchronize_user_dictionary(
    data: dict[str, Any],
    flow: pd.DataFrame,
) -> pd.DataFrame:
    current = data.get("dic_users", pd.DataFrame())
    if not isinstance(current, pd.DataFrame):
        current = pd.DataFrame(columns=["Correo", "Cargo"])

    result = current.copy()
    for column in ["Correo", "Cargo"]:
        if column not in result.columns:
            result[column] = ""

    result = result.loc[:, ["Correo", "Cargo"]].copy()
    result["Correo"] = result["Correo"].map(strip_user)
    result["Cargo"] = result["Cargo"].map(clean_text)

    existing = {
        email_key(value)
        for value in result["Correo"].tolist()
        if strip_user(value)
    }

    flow_users = sorted(
        {
            strip_user(value)
            for column in LIB_COLS
            for value in flow[column].tolist()
            if strip_user(value)
            and strip_user(value) != LS_LABEL
        },
        key=str.casefold,
    )

    missing_rows = [
        {"Correo": user, "Cargo": ""}
        for user in flow_users
        if email_key(user) not in existing
    ]

    if missing_rows:
        result = pd.concat(
            [result, pd.DataFrame(missing_rows)],
            ignore_index=True,
        )

    return (
        result[result["Correo"].map(strip_user).ne("")]
        .drop_duplicates("Correo", keep="last")
        .sort_values(
            "Correo",
            key=lambda values: values.astype(str).str.casefold(),
            kind="stable",
        )
        .reset_index(drop=True)
    )


def synchronize_ceco_dictionary(
    data: dict[str, Any],
    flow: pd.DataFrame,
) -> pd.DataFrame:
    current = data.get("dic_ceco", pd.DataFrame())
    if not isinstance(current, pd.DataFrame):
        current = pd.DataFrame(
            columns=["CECO", "Planta", "Centro"]
        )

    result = current.copy()
    for column in ["CECO", "Planta", "Centro"]:
        if column not in result.columns:
            result[column] = ""

    result = result.loc[:, ["CECO", "Planta", "Centro"]].copy()
    for column in ["CECO", "Planta", "Centro"]:
        result[column] = result[column].map(clean_text)

    flow_cecos = (
        flow[["CECO", "Planta"]]
        .copy()
        .assign(
            CECO=lambda frame: frame["CECO"].map(clean_text),
            Planta=lambda frame: frame["Planta"].map(clean_text),
        )
    )
    flow_cecos = (
        flow_cecos[flow_cecos["CECO"].ne("")]
        .drop_duplicates("CECO", keep="last")
    )

    existing_index = {
        clean_text(row["CECO"]): index
        for index, row in result.iterrows()
        if clean_text(row["CECO"])
    }

    new_rows: list[dict[str, str]] = []

    for _, row in flow_cecos.iterrows():
        ceco = clean_text(row["CECO"])
        planta = clean_text(row["Planta"])

        if ceco in existing_index:
            index = existing_index[ceco]
            if not clean_text(result.at[index, "Planta"]) and planta:
                result.at[index, "Planta"] = planta
        else:
            new_rows.append({
                "CECO": ceco,
                "Planta": planta,
                "Centro": "",
            })

    if new_rows:
        result = pd.concat(
            [result, pd.DataFrame(new_rows)],
            ignore_index=True,
        )

    return (
        result[result["CECO"].map(clean_text).ne("")]
        .drop_duplicates("CECO", keep="last")
        .sort_values("CECO", kind="stable")
        .reset_index(drop=True)
    )


def dataframe_to_csv_bytes(dataframe: pd.DataFrame) -> bytes:
    return clean_export_frame(dataframe).to_csv(
        index=False,
        sep=";",
        lineterminator="\n",
    ).encode("utf-8-sig")


def dataframe_to_parquet_bytes(dataframe: pd.DataFrame) -> bytes:
    engine = available_parquet_engine()

    if engine is None:
        raise ValueError(
            "Para descargar en Parquet debes instalar `pyarrow` "
            "o `fastparquet` en el entorno."
        )

    output = BytesIO()
    clean_export_frame(dataframe).to_parquet(
        output,
        index=False,
        engine=engine,
    )
    return output.getvalue()


def dataframe_to_excel_bytes(
    dataframe: pd.DataFrame,
    sheet_name: str = "Datos",
) -> bytes:
    output = BytesIO()
    clean_frame = clean_export_frame(dataframe)

    try:
        with pd.ExcelWriter(
            output,
            engine="openpyxl",
        ) as writer:
            clean_frame.to_excel(
                writer,
                index=False,
                sheet_name=sheet_name[:31],
            )
            worksheet = writer.book[sheet_name[:31]]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.sheet_view.showGridLines = False

            for column_cells in worksheet.columns:
                values = [
                    "" if cell.value is None else str(cell.value)
                    for cell in list(column_cells)[:250]
                ]
                width = min(
                    max(
                        max((len(value) for value in values), default=8) + 2,
                        12,
                    ),
                    45,
                )
                worksheet.column_dimensions[
                    column_cells[0].column_letter
                ].width = width
    except Exception as error:
        raise ValueError(
            "No fue posible generar el archivo Excel."
        ) from error

    return output.getvalue()


def archive_filename(
    role: str,
    extension: str,
    modification_timestamp: str,
) -> str:
    if role == "dic_ceco":
        stem = "Diccionario_CECO_Plantas"
    elif role == "dic_users":
        stem = "Diccionario_Usuarios_Cargos"
    else:
        level = int(role.rsplit("_", 1)[-1])
        stem = f"Liberador_{level}_Compra_Directa_ENAEX"

    return (
        f"{stem}_MODIFICADO_"
        f"{modification_timestamp}{extension}"
    )


def build_seven_files_archive(
    data: dict[str, Any],
    flow: pd.DataFrame,
    history: list[dict[str, Any]],
    export_format: str,
    modification_timestamp: str | None = None,
) -> tuple[
    bytes,
    dict[int, pd.DataFrame],
    pd.DataFrame,
    pd.DataFrame,
]:
    if export_format not in {"parquet", "csv", "excel"}:
        raise ValueError("Formato de descarga no válido.")

    timestamp = (
        modification_timestamp
        or chile_modification_timestamp()
    )

    original_frames = data.get("liberadores", {})
    if not isinstance(original_frames, dict):
        raise ValueError(
            "No se encontraron los cinco liberadores originales."
        )

    updated_frames = flow_to_level_frames(
        flow,
        original_frames,
    )
    updated_users = synchronize_user_dictionary(data, flow)
    updated_cecos = synchronize_ceco_dictionary(data, flow)

    logical_files: dict[str, pd.DataFrame] = {
        **{
            f"liberador_{level}": updated_frames[level]
            for level in LEVELS
        },
        "dic_ceco": updated_cecos,
        "dic_users": updated_users,
    }

    extension = {
        "parquet": ".parquet",
        "csv": ".csv",
        "excel": ".xlsx",
    }[export_format]

    zip_output = BytesIO()
    with zipfile.ZipFile(
        zip_output,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
    ) as archive:
        for role, dataframe in logical_files.items():
            if export_format == "parquet":
                content = dataframe_to_parquet_bytes(dataframe)
            elif export_format == "excel":
                content = dataframe_to_excel_bytes(
                    dataframe,
                    sheet_name=(
                        "CECO_Plantas"
                        if role == "dic_ceco"
                        else "Usuarios_Cargos"
                        if role == "dic_users"
                        else role.replace("_", " ").title()
                    ),
                )
            else:
                content = dataframe_to_csv_bytes(dataframe)
            archive.writestr(
                archive_filename(
                    role,
                    extension,
                    timestamp,
                ),
                content,
            )

        changes_columns = [
            "FechaHora",
            "Usuario",
            "CECO",
            "Desde",
            "Hasta",
            "TipoDoc",
            "Campo",
            "ValorAntes",
            "ValorDespues",
            "Nota",
        ]
        changes_dataframe = pd.DataFrame(
            history,
            columns=changes_columns,
        )
        archive.writestr(
            f"Cambios_MODIFICACION_{timestamp}.csv",
            dataframe_to_csv_bytes(changes_dataframe),
        )

    zip_bytes = zip_output.getvalue()

    # Garantiza la entrega contractual:
    # 5 liberadores + 2 diccionarios + 1 CSV de cambios.
    with zipfile.ZipFile(BytesIO(zip_bytes), mode="r") as validation_zip:
        archive_names = validation_zip.namelist()

    if len(archive_names) != 8:
        raise ValueError(
            "La exportación no generó los ocho archivos esperados."
        )

    expected_change_name = (
        f"Cambios_MODIFICACION_{timestamp}.csv"
    )
    if expected_change_name not in archive_names:
        raise ValueError(
            "No se generó correctamente el archivo CSV de cambios."
        )

    return (
        zip_bytes,
        updated_frames,
        updated_cecos,
        updated_users,
    )


def download_name_for_format(
    export_format: str,
    modification_timestamp: str,
) -> str:
    return (
        f"VERSION_MODIFICADA_"
        f"LIBERADORES_Y_DICCIONARIOS_"
        f"{export_format.upper()}_"
        f"{modification_timestamp}.zip"
    )


def refresh_download(
    file_name: Any,
    file_bytes: Any,
) -> None:
    data = st.session_state.get(SESSION_DATA_KEY)
    if not isinstance(data, dict):
        raise ValueError("No existe una versión activa.")

    flow = get_working_flow()
    history = list(
        st.session_state.get(SESSION_HISTORY_KEY, [])
    )
    timestamp = chile_modification_timestamp()

    (
        csv_zip,
        updated_frames,
        updated_cecos,
        updated_users,
    ) = build_seven_files_archive(
        data=data,
        flow=flow,
        history=history,
        export_format="csv",
        modification_timestamp=timestamp,
    )

    updated_data = dict(data)
    updated_data["liberadores"] = updated_frames
    for level in LEVELS:
        updated_data[f"liberador_{level}"] = updated_frames[level]
    updated_data["dic_ceco"] = updated_cecos
    updated_data["dic_users"] = updated_users

    st.session_state[SESSION_DATA_KEY] = updated_data
    st.session_state[SESSION_DOWNLOAD_CSV_KEY] = csv_zip
    st.session_state[SESSION_DOWNLOAD_CSV_NAME_KEY] = (
        download_name_for_format("csv", timestamp)
    )

    excel_zip, _, _, _ = build_seven_files_archive(
        data=updated_data,
        flow=flow,
        history=history,
        export_format="excel",
        modification_timestamp=timestamp,
    )
    st.session_state[SESSION_DOWNLOAD_EXCEL_KEY] = excel_zip
    st.session_state[SESSION_DOWNLOAD_EXCEL_NAME_KEY] = (
        download_name_for_format("excel", timestamp)
    )

    parquet_engine = available_parquet_engine()
    if parquet_engine is not None:
        parquet_zip, _, _, _ = build_seven_files_archive(
            data=updated_data,
            flow=flow,
            history=history,
            export_format="parquet",
            modification_timestamp=timestamp,
        )
        parquet_name = download_name_for_format(
            "parquet",
            timestamp,
        )
        st.session_state[SESSION_DOWNLOAD_PARQUET_KEY] = (
            parquet_zip
        )
        st.session_state[
            SESSION_DOWNLOAD_PARQUET_NAME_KEY
        ] = parquet_name
        st.session_state[SESSION_DOWNLOAD_KEY] = parquet_zip
        st.session_state[SESSION_DOWNLOAD_NAME_KEY] = parquet_name
    else:
        st.session_state.pop(
            SESSION_DOWNLOAD_PARQUET_KEY,
            None,
        )
        st.session_state.pop(
            SESSION_DOWNLOAD_PARQUET_NAME_KEY,
            None,
        )
        st.session_state[SESSION_DOWNLOAD_KEY] = csv_zip
        st.session_state[SESSION_DOWNLOAD_NAME_KEY] = (
            st.session_state[SESSION_DOWNLOAD_CSV_NAME_KEY]
        )




def render_download_selector(
    key_prefix: str,
) -> None:
    parquet_zip = st.session_state.get(
        SESSION_DOWNLOAD_PARQUET_KEY
    )
    parquet_name = st.session_state.get(
        SESSION_DOWNLOAD_PARQUET_NAME_KEY
    )
    csv_zip = st.session_state.get(
        SESSION_DOWNLOAD_CSV_KEY
    )
    csv_name = st.session_state.get(
        SESSION_DOWNLOAD_CSV_NAME_KEY
    )
    excel_zip = st.session_state.get(
        SESSION_DOWNLOAD_EXCEL_KEY
    )
    excel_name = st.session_state.get(
        SESSION_DOWNLOAD_EXCEL_NAME_KEY
    )

    if not any([parquet_zip, csv_zip, excel_zip]):
        return

    st.markdown("---")
    st.subheader("Descargar versión modificada")
    st.caption(
        "El ZIP contiene cinco liberadores, dos diccionarios "
        "actualizados y un archivo CSV de cambios. Todos los nombres "
        "incluyen la fecha y hora de modificación de Santiago."
    )

    selected_format = st.radio(
        "¿En qué formato deseas descargar los siete archivos principales?",
        options=["parquet", "csv", "excel"],
        index=0,
        format_func=lambda value: {
            "parquet": "Parquet — predeterminado",
            "csv": "CSV — separador punto y coma",
            "excel": "Excel — formato XLSX",
        }[value],
        key=f"{key_prefix}_format_v03",
    )

    payloads = {
        "parquet": (parquet_zip, parquet_name),
        "csv": (csv_zip, csv_name),
        "excel": (excel_zip, excel_name),
    }
    download_data, download_name = payloads[selected_format]

    if selected_format == "parquet" and not download_data:
        st.warning(
            "Parquet no está disponible en este entorno. "
            "Instala `pyarrow` o selecciona CSV/Excel."
        )
        return

    if not download_data or not download_name:
        st.error(
            f"No fue posible preparar la descarga en "
            f"{selected_format.upper()}."
        )
        return

    st.download_button(
        "⬇️ Descargar ZIP con 8 archivos",
        data=download_data,
        file_name=download_name,
        mime="application/zip",
        type="primary",
        use_container_width=True,
        key=f"{key_prefix}_{selected_format}_download_v03",
    )



# ============================================================
# REEMPLAZO GLOBAL DE PERSONA
# ============================================================

def render_global_replacement(
    data: dict[str, pd.DataFrame],
    file_name: str,
    file_bytes: bytes,
    actor: str,
    reason: str,
) -> None:
    """Reemplaza una persona en todas sus apariciones de la base."""
    flow = get_working_flow()
    users = unique_users(flow)

    question(
        2,
        "¿Qué persona se va de la empresa?",
        (
            "Selecciona la persona que debe ser reemplazada. "
            "Antes de guardar podrás revisar todos los CECO afectados."
        ),
    )

    if not users:
        st.warning("No se encontraron correos de liberadores en la base activa.")
        return

    old_user = st.selectbox(
        "Persona que sale",
        options=users,
        format_func=lambda value: display_user(value, data),
        key="global_old_user_v05",
    )

    occurrences = occurrences_of_person(flow, old_user)
    special_occurrences = special_occurrences_of_person(
        data,
        old_user,
    )

    if occurrences.empty and special_occurrences.empty:
        st.warning("La persona seleccionada no tiene apariciones en la base.")
        return

    affected_cecos = (
        int(occurrences["CECO"].nunique())
        if not occurrences.empty
        else 0
    )
    affected_rows = (
        int(occurrences["_ID_FILA"].nunique())
        if not occurrences.empty
        else 0
    )
    affected_positions = int(len(occurrences))
    affected_special = int(len(special_occurrences))

    metric_cols = st.columns(4)
    metrics = [
        ("CECO donde participa", affected_cecos),
        ("Filas CECO", affected_rows),
        ("Posiciones CECO", affected_positions),
        ("Reglas sin CECO", affected_special),
    ]

    for column, (label, value) in zip(metric_cols, metrics):
        with column:
            st.markdown(
                compact_html(
                    f"""
                    <div class="metric-card">
                        <div class="metric-label">{escape(label)}</div>
                        <div class="metric-value">{value}</div>
                    </div>
                    """
                ),
                unsafe_allow_html=True,
            )

    question(
        3,
        "¿Dónde participa actualmente?",
        (
            "Revisa todos los CECO, tramos, tipos y posiciones "
            "antes de seleccionar a la nueva persona."
        ),
    )

    occurrence_view = (
        occurrences[
            [
                "CECO",
                "Planta",
                "Desde",
                "Hasta",
                "TipoDoc",
                "Campo",
                "ValorAntes",
            ]
        ].rename(
            columns={
                "Campo": "Posición",
                "ValorAntes": "Persona actual",
            }
        )
        if not occurrences.empty
        else pd.DataFrame(
            columns=[
                "CECO",
                "Planta",
                "Desde",
                "Hasta",
                "TipoDoc",
                "Posición",
                "Persona actual",
            ]
        )
    )

    def style_occurrence_row(row: pd.Series) -> list[str]:
        doc = clean_text(row.get("TipoDoc")).upper()
        if doc == "AZNB":
            style = "background-color:#FFF1F0;color:#7A271A;"
        elif doc == "AZSR":
            style = "background-color:#EFF8FF;color:#1849A9;"
        else:
            style = ""
        return [style] * len(row)

    styled_occurrences = (
        occurrence_view.style
        .apply(style_occurrence_row, axis=1)
        .format(
            {
                "Desde": lambda value: fmt_bound(value),
                "Hasta": lambda value: fmt_bound(value),
            }
        )
    )

    if not occurrence_view.empty:
        st.dataframe(
            styled_occurrences,
            use_container_width=True,
            hide_index=True,
            height=min(
                600,
                max(260, 36 * (len(occurrence_view) + 2)),
            ),
        )

    if not special_occurrences.empty:
        st.markdown("#### Reglas especiales sin CECO explícito")
        st.warning(
            "Estas filas no aparecen en el flujo consolidado por CECO, "
            "pero también serán actualizadas por el reemplazo global."
        )
        st.dataframe(
            special_occurrences.style.apply(
                lambda row: [
                    "background-color:#FEF3C7;"
                    "color:#92400E;"
                    "font-weight:700;"
                ] * len(row),
                axis=1,
            ),
            use_container_width=True,
            hide_index=True,
        )

    question(
        4,
        "¿Quién será la nueva persona?",
        (
            "Selecciona un usuario existente o escribe un correo nuevo. "
            "El reemplazo se aplicará en todas las apariciones mostradas."
        ),
    )

    replacement_source = st.radio(
        "Origen de la nueva persona",
        options=["existing", "new"],
        format_func=lambda value: (
            "Seleccionar usuario existente"
            if value == "existing"
            else "Escribir correo nuevo"
        ),
        horizontal=True,
        key="global_replacement_source_v05",
    )

    new_user = ""

    if replacement_source == "existing":
        candidates = [
            user
            for user in users
            if email_key(user) != email_key(old_user)
        ]

        if candidates:
            new_user = st.selectbox(
                "Nueva persona",
                options=candidates,
                format_func=lambda value: display_user(value, data),
                key="global_new_existing_user_v05",
            )
        else:
            st.warning(
                "No hay otra persona disponible. "
                "Selecciona Escribir correo nuevo."
            )
    else:
        new_user = strip_user(
            st.text_input(
                "Correo de la nueva persona",
                placeholder="nombre.apellido@enaex.com",
                key="global_new_email_v05",
            )
        )

    question(
        5,
        "¿Confirmas el reemplazo global?",
        (
            "Se reemplazará únicamente a la persona seleccionada; "
            "los demás liberadores y el orden del flujo permanecerán sin cambios."
        ),
    )

    if new_user:
        st.info(
            f"Se reemplazará **{display_user(old_user, data)}** por "
            f"**{display_user(new_user, data)}** en "
            f"**{affected_cecos} CECO**, **{affected_rows} filas CECO**, "
            f"**{affected_positions} posiciones** y "
            f"**{affected_special} reglas especiales sin CECO**."
        )

    confirmation = st.checkbox(
        (
            "Confirmo que la persona seleccionada debe ser reemplazada "
            "en toda la base."
        ),
        value=False,
        key="global_confirm_v05",
    )

    global_ready = bool(
        new_user
        and confirmation
        and clean_text(actor)
        and clean_text(reason)
        and is_valid_email(new_user)
    )

    if new_user and not is_valid_email(new_user):
        st.error("El correo de la nueva persona no tiene un formato válido.")

    if not clean_text(actor) or not clean_text(reason):
        st.warning(
            "Para realizar un reemplazo global debes indicar "
            "quién modifica y el motivo."
        )

    apply_clicked = st.button(
        "🔁 Aplicar reemplazo global y preparar 8 archivos",
        type="primary",
        use_container_width=True,
        disabled=not global_ready,
        key="global_apply_v06",
    )

    if apply_clicked:
        old_key = email_key(old_user)
        replacement = strip_user(new_user)

        if not replacement:
            st.error("Indica la nueva persona.")
            return

        if email_key(replacement) == old_key:
            st.error("La nueva persona debe ser distinta de la persona que sale.")
            return

        updated = flow.copy(deep=True)
        timestamp = datetime.now(CHILE_TZ).strftime("%Y-%m-%d %H:%M:%S")
        changes: list[dict[str, Any]] = []

        for row_index, row in updated.iterrows():
            for column in LIB_COLS:
                current = strip_user(row[column])

                if email_key(current) != old_key:
                    continue

                updated.at[row_index, column] = replacement

                changes.append(
                    {
                        "FechaHora": timestamp,
                        "Usuario": actor or "anonimo",
                        "CECO": row["CECO"],
                        "Desde": row["Desde"],
                        "Hasta": row["Hasta"],
                        "TipoDoc": row["TipoDoc"],
                        "Campo": column,
                        "ValorAntes": current,
                        "ValorDespues": replacement,
                        "Nota": (
                            reason
                            or "Reemplazo global por salida de la empresa"
                        ),
                    }
                )

        current_data = st.session_state.get(
            SESSION_DATA_KEY,
            data,
        )
        if isinstance(current_data, dict):
            updated_data, special_changes = (
                replace_user_in_special_rules(
                    current_data,
                    old_user,
                    replacement,
                    actor,
                    reason,
                    timestamp,
                )
            )
            st.session_state[SESSION_DATA_KEY] = updated_data
            changes.extend(special_changes)

        set_working_flow(updated)

        history = list(st.session_state.get(SESSION_HISTORY_KEY, []))
        history.extend(changes)
        st.session_state[SESSION_HISTORY_KEY] = history

        try:
            refresh_download(file_name, file_bytes)
            st.success(
                f"Reemplazo global completado: "
                f"**{len(changes)} apariciones** actualizadas, "
                f"incluidas **{affected_special} reglas sin CECO**, "
                f"en **{affected_cecos} CECO**."
            )
            st.toast("Reemplazo global guardado.", icon="✅")
        except ValueError as error:
            st.error(str(error))

    render_download_selector("global")

    history = st.session_state.get(SESSION_HISTORY_KEY, [])
    if history:
        with st.expander(
            f"Historial de esta sesión ({len(history)} cambios)",
            expanded=False,
        ):
            st.dataframe(
                pd.DataFrame(history),
                use_container_width=True,
                hide_index=True,
            )



# ============================================================
# REEMPLAZO DE USUARIO EN CECO COMPLETO
# ============================================================

def render_ceco_user_replacement(
    data: dict[str, pd.DataFrame],
    file_name: str,
    file_bytes: bytes,
    actor: str,
    reason: str,
) -> None:
    """
    Reemplaza un usuario en todas las posiciones de un CECO, sin filtrar
    por rango. Puede incluir el CECO gemelo EMTS y uno o ambos tipos.
    """
    flow = get_working_flow()

    question(
        2,
        "¿En qué CECO quieres reemplazar al usuario?",
        (
            "Selecciona el CECO principal. Si existe un gemelo EMTS, "
            "la aplicación lo detectará y permitirá incluirlo."
        ),
    )

    ceco_map = (
        flow[["CECO", "Planta"]]
        .drop_duplicates()
        .sort_values(["Planta", "CECO"], kind="stable")
        .groupby("CECO", as_index=False)
        .first()
    )

    if ceco_map.empty:
        st.warning("No existen CECO disponibles en la base activa.")
        return

    plant_by_ceco = dict(zip(ceco_map["CECO"], ceco_map["Planta"]))

    selected_ceco = st.selectbox(
        "CECO principal",
        options=ceco_map["CECO"].tolist(),
        format_func=lambda value: (
            f"{value} | {plant_by_ceco.get(value, '')}"
            if plant_by_ceco.get(value, "")
            else value
        ),
        key="ceco_replace_selected_v01",
    )

    twin_records = find_twin_cecos(flow, selected_ceco)
    include_twin = False

    if twin_records:
        twin_text = ", ".join(
            f"{item['ceco']} | {item['planta']}"
            for item in twin_records
        )
        st.info(f"CECO gemelo detectado: **{twin_text}**.")

        include_twin = st.checkbox(
            "Aplicar también al CECO gemelo EMTS",
            value=True,
            key="ceco_replace_include_twin_v01",
            help=(
                "Está activado por defecto para mantener sincronizados "
                "el CECO principal y su gemelo EMTS."
            ),
        )

    target_cecos = [selected_ceco]

    if include_twin:
        target_cecos.extend(
            item["ceco"]
            for item in twin_records
        )

    question(
        3,
        "¿En qué tipo de documento se aplicará?",
        (
            "Puedes reemplazar al usuario en Material, Servicio "
            "o en ambos tipos. El rango no limita esta operación."
        ),
    )

    available_docs = [
        doc
        for doc in ["AZNB", "AZSR"]
        if not flow[
            flow["CECO"].isin(target_cecos)
            & flow["TipoDoc"].eq(doc)
        ].empty
    ]

    if not available_docs:
        st.warning(
            "Los CECO seleccionados no contienen reglas AZNB ni AZSR."
        )
        return

    doc_options = list(available_docs)

    if {"AZNB", "AZSR"}.issubset(set(available_docs)):
        doc_options.insert(0, "AMBOS")

    selected_doc = st.selectbox(
        "Tipo de documento",
        options=doc_options,
        format_func=lambda value: DOC_LABEL[value],
        key="ceco_replace_doc_v01",
    )

    target_docs = (
        ["AZNB", "AZSR"]
        if selected_doc == "AMBOS"
        else [selected_doc]
    )

    scope_rows = flow[
        flow["CECO"].isin(target_cecos)
        & flow["TipoDoc"].isin(target_docs)
    ].copy()

    if scope_rows.empty:
        st.warning("No existen filas dentro del alcance seleccionado.")
        return

    st.markdown("#### Tabla completa del alcance")
    st.caption(
        (
            f"Se muestran **{len(scope_rows):,} filas** de "
            f"**{scope_rows['CECO'].nunique():,} CECO**, sin filtrar por usuario. "
            "Esta primera tabla muestra el alcance general. Después de seleccionar "
            "a la persona, aparecerá nuevamente completa con las filas "
            "afectadas resaltadas para comparar contra las demás."
        ).replace(",", ".")
    )

    complete_scope = scope_rows.copy()
    complete_scope["_TipoOrden"] = complete_scope["TipoDoc"].map(
        {"AZNB": 0, "AZSR": 1}
    ).fillna(99)
    complete_scope["_DesdeOrden"] = complete_scope["Desde"].map(
        lambda value: parse_bound(value, low=True)
    )
    complete_scope = (
        complete_scope.sort_values(
            ["Planta", "CECO", "_TipoOrden", "_DesdeOrden", "_ID_FILA"],
            kind="stable",
        )
        .drop(columns=["_TipoOrden", "_DesdeOrden"])
    )

    st.dataframe(
        style_flow_table(complete_scope),
        use_container_width=True,
        hide_index=True,
        height=min(
            700,
            max(300, 36 * (len(complete_scope) + 2)),
        ),
    )

    question(
        4,
        "¿Qué usuario antiguo quieres reemplazar?",
        (
            "La lista contiene solo usuarios que aparecen en los CECO "
            "y tipos seleccionados. Las posiciones afectadas se marcarán "
            "en amarillo en la vista previa."
        ),
    )

    scoped_users = unique_users(scope_rows)

    if not scoped_users:
        st.warning(
            "No se encontraron usuarios reemplazables dentro del alcance."
        )
        return

    old_user = st.selectbox(
        "Usuario antiguo",
        options=scoped_users,
        format_func=lambda value: display_user(value, data),
        key="ceco_replace_old_user_v01",
    )

    old_key = email_key(old_user)

    occurrences = occurrences_of_person(
        scope_rows,
        old_user,
    )

    if occurrences.empty:
        st.warning(
            "El usuario seleccionado no tiene apariciones dentro del alcance."
        )
        return

    affected_row_ids = (
        occurrences["_ID_FILA"]
        .drop_duplicates()
        .astype(int)
        .tolist()
    )
    affected_rows = scope_rows[
        scope_rows["_ID_FILA"].isin(affected_row_ids)
    ].copy()

    affected_cecos = int(occurrences["CECO"].nunique())
    affected_ranges = int(
        occurrences[
            ["CECO", "TipoDoc", "Desde", "Hasta"]
        ]
        .drop_duplicates()
        .shape[0]
    )
    affected_positions = int(len(occurrences))

    metric_columns = st.columns(4)
    metrics = [
        ("CECO afectados", affected_cecos),
        ("Filas afectadas", len(affected_rows)),
        ("Tramos afectados", affected_ranges),
        ("Posiciones", affected_positions),
    ]

    for column, (label, value) in zip(metric_columns, metrics):
        with column:
            st.markdown(
                compact_html(
                    f"""
                    <div class="metric-card">
                        <div class="metric-label">{escape(label)}</div>
                        <div class="metric-value">{value}</div>
                    </div>
                    """
                ),
                unsafe_allow_html=True,
            )

    st.markdown("#### Vista previa de posiciones afectadas")

    st.markdown(
        compact_html(
            f"""
            <div class="replacement-note">
                Se muestra la tabla completa del alcance para facilitar la
                comparación. Las filas amarillas contienen al menos una posición
                que será modificada. Las celdas con borde amarillo intenso contienen
                {escape(display_user(old_user, data))} y serán reemplazadas.
            </div>
            """
        ),
        unsafe_allow_html=True,
    )

    visible_columns = [
        "Estado", "CECO", "Planta", "Desde", "Hasta", "TipoDoc",
        "Lib1", "Lib2", "Lib3", "Lib4", "Lib5",
    ]

    preview = complete_scope.copy()
    preview["Estado"] = preview["_ID_FILA"].map(
        lambda row_id: (
            "AFECTADA"
            if int(row_id) in set(affected_row_ids)
            else "REFERENCIA"
        )
    )
    preview = preview.loc[:, visible_columns].copy()

    def highlight_affected_rows(row: pd.Series) -> list[str]:
        if clean_text(row.get("Estado", "")) == "AFECTADA":
            style = (
                "background-color:#FEF3C7;"
                "color:#78350F;"
            )
            return [style] * len(row)

        return [""] * len(row)

    def highlight_replacement_cells(column: pd.Series) -> list[str]:
        return [
            (
                "background-color:#FDE047;"
                "color:#713F12;"
                "font-weight:850;"
                "border:2px solid #F59E0B;"
            )
            if email_key(value) == old_key
            else ""
            for value in column
        ]

    def highlight_status(column: pd.Series) -> list[str]:
        return [
            (
                "background-color:#FDE68A;"
                "color:#78350F;"
                "font-weight:850;"
                "text-align:center;"
            )
            if clean_text(value) == "AFECTADA"
            else (
                "background-color:#F8FAFC;"
                "color:#64748B;"
                "text-align:center;"
            )
            for value in column
        ]

    styled_preview = (
        preview.style
        .apply(highlight_affected_rows, axis=1)
        .apply(
            highlight_replacement_cells,
            subset=LIB_COLS,
            axis=0,
        )
        .apply(
            highlight_status,
            subset=["Estado"],
            axis=0,
        )
        .format(
            {
                "Desde": lambda value: fmt_bound(value),
                "Hasta": lambda value: fmt_bound(value),
            }
        )
        .set_properties(
            subset=["Estado"],
            **{"font-size": "11px"},
        )
    )

    st.caption(
        (
            f"Comparación completa: **{len(preview):,} filas visibles** · "
            f"**{len(affected_rows):,} afectadas** · "
            f"**{len(preview) - len(affected_rows):,} de referencia**."
        ).replace(",", ".")
    )

    st.dataframe(
        styled_preview,
        use_container_width=True,
        hide_index=True,
        height=min(
            760,
            max(320, 38 * (len(preview) + 2)),
        ),
    )

    with st.expander(
        "Detalle exacto de posiciones que serán reemplazadas",
        expanded=True,
    ):
        occurrence_view = occurrences[
            [
                "CECO",
                "Planta",
                "Desde",
                "Hasta",
                "TipoDoc",
                "Campo",
                "ValorAntes",
            ]
        ].rename(
            columns={
                "Campo": "Posición",
                "ValorAntes": "Usuario antiguo",
            }
        )

        occurrence_view["Desde"] = occurrence_view["Desde"].map(fmt_bound)
        occurrence_view["Hasta"] = occurrence_view["Hasta"].map(fmt_bound)

        st.dataframe(
            occurrence_view,
            use_container_width=True,
            hide_index=True,
            height=min(
                650,
                max(240, 36 * (len(occurrence_view) + 2)),
            ),
        )

    question(
        5,
        "¿Qué usuario nuevo debe reemplazarlo?",
        (
            "Selecciona un usuario existente o escribe un correo nuevo. "
            "Se conservarán las mismas posiciones Lib1 a Lib5."
        ),
    )

    replacement_source = st.radio(
        "Origen del usuario nuevo",
        options=["existing", "new"],
        format_func=lambda value: (
            "Seleccionar usuario existente"
            if value == "existing"
            else "Escribir correo nuevo"
        ),
        horizontal=True,
        key="ceco_replace_source_v01",
    )

    new_user = ""

    if replacement_source == "existing":
        all_users = unique_users(flow)
        candidates = [
            user
            for user in all_users
            if email_key(user) != old_key
        ]

        if candidates:
            new_user = st.selectbox(
                "Usuario nuevo",
                options=candidates,
                format_func=lambda value: display_user(value, data),
                key="ceco_replace_new_existing_v01",
            )
        else:
            st.warning(
                "No existen otros usuarios disponibles. "
                "Selecciona Escribir correo nuevo."
            )
    else:
        new_user = strip_user(
            st.text_input(
                "Correo del usuario nuevo",
                placeholder="nombre.apellido@enaex.com",
                key="ceco_replace_new_text_v01",
            )
        )

    if new_user:
        duplicated_rows: list[str] = []

        for _, row in affected_rows.iterrows():
            row_users = [
                email_key(row.get(column, ""))
                for column in LIB_COLS
                if email_key(row.get(column, ""))
            ]

            if (
                email_key(new_user) in row_users
                and email_key(new_user) != old_key
            ):
                duplicated_rows.append(
                    (
                        f"{row['CECO']} / {row['TipoDoc']} / "
                        f"{fmt_bound(row['Desde'])}–{fmt_bound(row['Hasta'])}"
                    )
                )

        if duplicated_rows:
            st.error(
                "El usuario nuevo ya aparece en algunas filas afectadas. "
                "El reemplazo generaría duplicados en: "
                + "; ".join(duplicated_rows[:8])
                + ("…" if len(duplicated_rows) > 8 else "")
            )

    question(
        6,
        "¿Confirmas el reemplazo?",
        (
            "La operación reemplazará solo las celdas destacadas, "
            "manteniendo rangos, tipos, posiciones y demás usuarios."
        ),
    )

    if new_user:
        st.info(
            f"Se reemplazará **{display_user(old_user, data)}** por "
            f"**{display_user(new_user, data)}** en "
            f"**{affected_positions} posiciones**, "
            f"**{len(affected_rows)} filas** y "
            f"**{affected_cecos} CECO**."
        )

    duplicate_conflict = False

    if new_user:
        for _, row in affected_rows.iterrows():
            row_keys = [
                email_key(row.get(column, ""))
                for column in LIB_COLS
                if email_key(row.get(column, ""))
            ]
            if (
                email_key(new_user) in row_keys
                and email_key(new_user) != old_key
            ):
                duplicate_conflict = True
                break

    confirmation = st.checkbox(
        (
            "Confirmo el reemplazo en todos los rangos y posiciones "
            "destacados."
        ),
        value=False,
        key="ceco_replace_confirm_v01",
    )

    ready = bool(
        new_user
        and is_valid_email(new_user)
        and email_key(new_user) != old_key
        and not duplicate_conflict
        and confirmation
        and clean_text(actor)
        and clean_text(reason)
    )

    if new_user and not is_valid_email(new_user):
        st.error("El correo del usuario nuevo no tiene un formato válido.")

    if new_user and email_key(new_user) == old_key:
        st.error("El usuario nuevo debe ser distinto del usuario antiguo.")

    if not clean_text(actor) or not clean_text(reason):
        st.warning(
            "Completa ¿Quién modifica? y ¿Por qué? para habilitar el guardado."
        )

    apply_clicked = st.button(
        "🔁 Aplicar reemplazo y preparar Excel",
        type="primary",
        use_container_width=True,
        disabled=not ready,
        key="ceco_replace_apply_v01",
    )

    if apply_clicked:
        updated = flow.copy(deep=True)
        timestamp = datetime.now(CHILE_TZ).strftime("%Y-%m-%d %H:%M:%S")
        replacement = strip_user(new_user)
        changes: list[dict[str, Any]] = []

        target_index = updated[
            updated["_ID_FILA"].isin(affected_row_ids)
        ].index

        for row_index in target_index:
            row = updated.loc[row_index]

            for column in LIB_COLS:
                current = strip_user(row.get(column, ""))

                if email_key(current) != old_key:
                    continue

                updated.at[row_index, column] = replacement

                changes.append(
                    {
                        "FechaHora": timestamp,
                        "Usuario": actor or "anonimo",
                        "CECO": row["CECO"],
                        "Desde": row["Desde"],
                        "Hasta": row["Hasta"],
                        "TipoDoc": row["TipoDoc"],
                        "Campo": column,
                        "ValorAntes": current,
                        "ValorDespues": replacement,
                        "Nota": (
                            f"{reason}. Reemplazo completo por CECO, "
                            f"independiente del rango"
                            + (
                                ", incluyendo gemelo EMTS"
                                if include_twin
                                else ""
                            )
                            + (
                                ", ambos tipos de documento"
                                if selected_doc == "AMBOS"
                                else ""
                            )
                        ),
                    }
                )

        set_working_flow(updated)

        history = list(
            st.session_state.get(
                SESSION_HISTORY_KEY,
                [],
            )
        )
        history.extend(changes)
        st.session_state[SESSION_HISTORY_KEY] = history

        try:
            refresh_download(file_name, file_bytes)
            st.success(
                f"Reemplazo completado: **{len(changes)} posiciones** "
                f"actualizadas en **{affected_cecos} CECO**."
            )
            st.toast(
                "Reemplazo por CECO guardado.",
                icon="✅",
            )
        except ValueError as error:
            st.error(str(error))

    render_download_selector("ceco_replace")

    history = st.session_state.get(
        SESSION_HISTORY_KEY,
        [],
    )

    if history:
        with st.expander(
            f"Historial de esta sesión ({len(history)} cambios)",
            expanded=False,
        ):
            st.dataframe(
                pd.DataFrame(history),
                use_container_width=True,
                hide_index=True,
            )


# ============================================================
# INTERFAZ PASO A PASO
# ============================================================

def render_wizard(
    data: dict[str, pd.DataFrame],
    file_name: str,
    file_bytes: bytes,
) -> None:
    flow = get_working_flow()
    draft = get_draft()

    # --------------------------------------------------------
    # Datos del cambio
    # --------------------------------------------------------
    with st.expander("Identificación del cambio", expanded=True):
        actor_col, reason_col = st.columns(2)

        with actor_col:
            actor = st.text_input(
                "¿Quién modifica?",
                placeholder="nombre.apellido",
                key="mod_actor_v04",
            )

        with reason_col:
            reason = st.text_input(
                "¿Por qué?",
                placeholder="vacaciones, reorden, reemplazo...",
                key="mod_reason_v04",
            )

    # --------------------------------------------------------
    # 1. Escenario de negocio
    # --------------------------------------------------------
    question(
        1,
        "¿Qué situación necesitas resolver?",
        (
            "Selecciona el escenario más parecido a tu necesidad. "
            "La aplicación limitará las opciones para evitar errores."
        ),
    )

    scenario = st.selectbox(
        "1. ¿Qué situación necesitas resolver?",
        options=list(SCENARIO_LABEL),
        format_func=lambda value: SCENARIO_LABEL[value],
        key="mod_scenario_v07",
        help=(
            "Selecciona el escenario que mejor representa el cambio. "
            "Las acciones disponibles se ajustarán automáticamente."
        ),
    )
    st.info(SCENARIO_HELP[scenario])

    if scenario == "reemplazo_ceco":
        render_ceco_user_replacement(
            data=data,
            file_name=file_name,
            file_bytes=file_bytes,
            actor=actor,
            reason=reason,
        )
        return

    if scenario == "salida":
        if not clean_text(actor) or not clean_text(reason):
            st.warning(
                "Completa ¿Quién modifica? y ¿Por qué? antes de "
                "aplicar un reemplazo global."
            )
        render_global_replacement(
            data=data,
            file_name=file_name,
            file_bytes=file_bytes,
            actor=actor,
            reason=reason,
        )
        return

    # --------------------------------------------------------
    # 2. CECO y gemelo EMTS
    # --------------------------------------------------------
    question(
        2,
        "¿Qué CECO quieres modificar?",
        (
            "Selecciona el CECO. Si existe un CECO gemelo EMTS, "
            "podrás incluirlo en la misma modificación."
        ),
    )

    ceco_map = (
        flow[["CECO", "Planta"]]
        .drop_duplicates()
        .sort_values(["CECO", "Planta"], kind="stable")
        .groupby("CECO", as_index=False)
        .first()
    )
    plant_by_ceco = dict(zip(ceco_map["CECO"], ceco_map["Planta"]))

    selected_ceco = st.selectbox(
        "CECO",
        options=ceco_map["CECO"].tolist(),
        format_func=lambda value: (
            f"{value} | {plant_by_ceco.get(value, '')}"
            if plant_by_ceco.get(value, "")
            else value
        ),
        key="mod_ceco_v08",
    )

    twin_records = find_twin_cecos(flow, selected_ceco)
    include_twin = False

    if twin_records:
        twin_description = ", ".join(
            f"{item['planta']} ({item['ceco']})"
            for item in twin_records
        )
        st.info(
            f"Se detectó CECO gemelo para **{plant_by_ceco.get(selected_ceco, selected_ceco)}**: "
            f"**{twin_description}**."
        )
        include_twin = st.checkbox(
            "Aplicar también la modificación al CECO gemelo EMTS",
            value=False,
            key="mod_include_twin_v08",
            help=(
                "Al activarlo se modificarán las filas equivalentes del CECO "
                "seleccionado y de su gemelo, usando el mismo tipo y rango."
            ),
        )

    target_cecos = [selected_ceco]
    if include_twin:
        target_cecos.extend(item["ceco"] for item in twin_records)

    ceco_rows = flow[flow["CECO"].isin(target_cecos)].copy()

    st.dataframe(
        style_flow_table(ceco_rows),
        use_container_width=True,
        hide_index=True,
        height=min(520, max(250, 36 * (len(ceco_rows) + 2))),
    )

    # --------------------------------------------------------
    # 3. Tipo de documento
    # --------------------------------------------------------
    question(
        3,
        "¿Qué tipo quieres modificar?",
        "Selecciona Material, Servicio o Ambos.",
    )

    primary_rows = flow[flow["CECO"].eq(selected_ceco)].copy()
    available_docs = [
        doc
        for doc in ["AZNB", "AZSR"]
        if not primary_rows[primary_rows["TipoDoc"].eq(doc)].empty
    ]

    if not available_docs:
        st.error("El CECO no contiene reglas AZNB ni AZSR.")
        return

    doc_options = list(available_docs)
    if {"AZNB", "AZSR"}.issubset(set(available_docs)):
        doc_options.append("AMBOS")

    selected_doc = st.selectbox(
        "Tipo de documento",
        options=doc_options,
        format_func=lambda value: DOC_LABEL[value],
        key="mod_doc_v08",
        help=(
            "Ambos aplica la misma modificación a Material (AZNB) "
            "y Servicio (AZSR) para el rango seleccionado."
        ),
    )

    selected_docs = (
        ["AZNB", "AZSR"]
        if selected_doc == "AMBOS"
        else [selected_doc]
    )

    # --------------------------------------------------------
    # 4. Rango
    # --------------------------------------------------------
    question(
        4,
        "¿Qué rango quieres abrir?",
        (
            "Selecciona el tramo. La aplicación buscará el mismo rango "
            "en los tipos y CECO incluidos."
        ),
    )

    range_source = primary_rows[
        primary_rows["TipoDoc"].isin(selected_docs)
    ].copy()
    range_source["_DESDE"] = range_source["Desde"].map(
        lambda value: parse_bound(value, low=True)
    )
    range_source["_HASTA"] = range_source["Hasta"].map(
        lambda value: parse_bound(value, low=False)
    )
    range_options = (
        range_source[["_DESDE", "_HASTA", "Desde", "Hasta"]]
        .drop_duplicates(["_DESDE", "_HASTA"])
        .sort_values(["_DESDE", "_HASTA"], kind="stable")
        .reset_index(drop=True)
    )

    if range_options.empty:
        st.error("No se encontraron rangos para la selección actual.")
        return

    selected_range_index = st.selectbox(
        "Rango",
        options=range_options.index.tolist(),
        format_func=lambda index: (
            f"{fmt_bound(range_options.loc[index, 'Desde'])} – "
            f"{fmt_bound(range_options.loc[index, 'Hasta'])}"
        ),
        key="mod_range_v08",
    )

    selected_from = range_options.loc[selected_range_index, "Desde"]
    selected_until = range_options.loc[selected_range_index, "Hasta"]

    target_rows = flow[
        flow["CECO"].isin(target_cecos)
        & flow["TipoDoc"].isin(selected_docs)
        & same_range_mask(flow, selected_from, selected_until)
    ].copy()

    if target_rows.empty:
        st.error("No existen filas equivalentes para la selección realizada.")
        return

    expected_combinations = {
        (ceco, doc)
        for ceco in target_cecos
        for doc in selected_docs
    }
    found_combinations = set(
        target_rows[["CECO", "TipoDoc"]]
        .drop_duplicates()
        .itertuples(index=False, name=None)
    )
    missing_combinations = sorted(expected_combinations - found_combinations)

    if missing_combinations:
        missing_text = ", ".join(
            f"{ceco} / {DOC_LABEL.get(doc, doc)}"
            for ceco, doc in missing_combinations
        )
        st.warning(
            "No se encontró el mismo rango para todas las combinaciones. "
            f"No serán modificadas: **{missing_text}**."
        )

    target_rows = target_rows.sort_values(
        ["CECO", "TipoDoc", "_ID_FILA"],
        kind="stable",
    )
    target_row_ids = target_rows["_ID_FILA"].astype(int).tolist()

    primary_candidates = target_rows[
        target_rows["CECO"].eq(selected_ceco)
    ]
    selected_row = (
        primary_candidates.iloc[0]
        if not primary_candidates.empty
        else target_rows.iloc[0]
    )
    selected_row_id = int(selected_row["_ID_FILA"])

    current_identity = (
        tuple(target_cecos),
        tuple(selected_docs),
        float(parse_bound(selected_from, low=True)),
        float(parse_bound(selected_until, low=False)),
        tuple(target_row_ids),
    )
    draft_identity = (
        tuple(draft.get("target_cecos", [])),
        tuple(draft.get("target_docs", [])),
        draft.get("range_from"),
        draft.get("range_until"),
        tuple(draft.get("target_row_ids", [])),
    )

    if current_identity != draft_identity:
        libs = libs_from_row(selected_row)
        draft = default_draft()
        draft.update(
            {
                "ceco": selected_ceco,
                "doc": selected_doc,
                "row_id": selected_row_id,
                "target_row_ids": target_row_ids,
                "target_cecos": list(target_cecos),
                "target_docs": list(selected_docs),
                "range_from": float(parse_bound(selected_from, low=True)),
                "range_until": float(parse_bound(selected_until, low=False)),
                "libs_before": list(libs),
                "libs_after": list(libs),
            }
        )
        set_draft(draft)

    libs_before = list(draft["libs_before"])
    libs_after = list(draft["libs_after"])

    signatures = {
        row_flow_signature(row)
        for _, row in target_rows.iterrows()
    }
    if len(signatures) > 1:
        st.warning(
            "Las filas seleccionadas no tienen exactamente el mismo flujo actual. "
            "La edición usará como base el flujo del CECO principal y, al guardar, "
            "aplicará el resultado final a todas las filas indicadas."
        )

    st.info(
        f"Tramo activo: **{fmt_bound(selected_from)} – {fmt_bound(selected_until)}** · "
        f"**{len(target_rows)} fila(s)** · "
        f"**{len(set(target_rows['CECO']))} CECO** · "
        f"**{len(set(target_rows['TipoDoc']))} tipo(s)**."
    )

    with st.expander("Ver filas que serán modificadas", expanded=True):
        st.dataframe(
            style_flow_table(target_rows),
            use_container_width=True,
            hide_index=True,
            height=min(520, max(220, 36 * (len(target_rows) + 2))),
        )

    st.markdown(
        flow_html(
            libs_after,
            data,
            "Flujo que se aplicará a las filas seleccionadas",
        ),
        unsafe_allow_html=True,
    )

    # --------------------------------------------------------
    # 4. Pieza
    # --------------------------------------------------------
    question(
        5,
        "¿Qué pieza quieres modificar?",
        "Para Agregar no es necesario seleccionar una pieza.",
    )

    piece_options = list(range(len(libs_after)))
    selected_piece = None

    if piece_options:
        selected_piece = st.selectbox(
            "Pieza",
            options=piece_options,
            format_func=lambda index: (
                f"Liberador {index + 1}: "
                f"{display_user(libs_after[index], data)}"
            ),
            key="mod_piece_v04",
        )
    else:
        st.warning(
            "El flujo no tiene liberadores. Utiliza la acción Agregar."
        )

    # --------------------------------------------------------
    # 5. Acción
    # --------------------------------------------------------
    question(
        6,
        "¿Qué quieres hacer?",
        "La subpregunta cambia según la acción seleccionada.",
    )

    scenario_actions = {
        "ajuste": ["mover", "reemplazar", "eliminar", "agregar"],
        "temporal": ["reemplazar"],
        "orden": ["mover"],
        "dotacion": ["agregar", "eliminar"],
    }
    allowed_actions = scenario_actions.get(
        scenario,
        list(ACTION_LABEL),
    )

    action = st.selectbox(
        "Acción",
        options=allowed_actions,
        format_func=lambda value: ACTION_LABEL[value],
        key="mod_action_v06",
    )

    destination_index = None
    new_value = ""

    if action == "mover":
        if not piece_options:
            st.warning("No existen piezas para mover.")
        else:
            destination_index = st.selectbox(
                "¿A qué posición?",
                options=piece_options,
                format_func=lambda index: f"Liberador {index + 1}",
                key="mod_destination_v04",
            )

    elif action in {"reemplazar", "agregar"}:
        prompt = (
            "¿Qué entra?"
            if action == "reemplazar"
            else "¿Qué agregas?"
        )

        entry_mode = st.radio(
            prompt,
            options=["mail", "ls"],
            format_func=lambda value: (
                "Correo"
                if value == "mail"
                else "Liberador Servicios"
            ),
            horizontal=True,
            key="mod_entry_mode_v04",
        )

        if entry_mode == "ls":
            new_value = LS_LABEL
        else:
            users = unique_users(flow)
            selection_mode = st.radio(
                "Origen del correo",
                options=["existing", "new"],
                format_func=lambda value: (
                    "Seleccionar usuario existente"
                    if value == "existing"
                    else "Escribir correo nuevo"
                ),
                horizontal=True,
                key="mod_user_mode_v04",
            )

            if selection_mode == "existing" and users:
                new_value = st.selectbox(
                    "Correo",
                    options=users,
                    format_func=lambda value: display_user(value, data),
                    key="mod_existing_user_v04",
                )
            else:
                new_value = strip_user(
                    st.text_input(
                        "Correo",
                        placeholder="nombre.apellido@enaex.com",
                        key="mod_new_user_v04",
                    )
                )

    apply_clicked = st.button(
        "Aplicar respuesta",
        type="primary",
        use_container_width=True,
        key="mod_apply_answer_v04",
    )

    if apply_clicked:
        try:
            if (
                action in {"reemplazar", "agregar"}
                and new_value
                and not is_valid_email(new_value)
            ):
                raise ValueError(
                    "El correo ingresado no tiene un formato válido."
                )

            result, message, replaced_from, replaced_to = apply_action(
                action=action,
                libs=libs_after,
                selected_index=(
                    selected_piece
                    if action != "agregar"
                    else None
                ),
                destination_index=destination_index,
                new_value=strip_user(new_value),
            )

            draft["libs_after"] = result
            draft["last_message"] = message
            draft["replaced_from"] = replaced_from
            draft["replaced_to"] = replaced_to
            set_draft(draft)
            st.rerun()

        except ValueError as error:
            st.error(str(error))

    if draft.get("last_message"):
        st.success(draft["last_message"])

    # --------------------------------------------------------
    # Vista previa
    # --------------------------------------------------------
    st.markdown("### Vista previa")
    render_comparison(
        before=libs_before,
        after=libs_after,
        data=data,
    )

    col_undo, col_restart = st.columns(2)

    with col_undo:
        if st.button(
            "↩️ Retroceder último borrador",
            use_container_width=True,
            disabled=libs_after == libs_before,
            key="mod_undo_v04",
        ):
            draft["libs_after"] = list(libs_before)
            draft["replaced_from"] = ""
            draft["replaced_to"] = ""
            draft["last_message"] = "Se restauró el flujo original del tramo."
            set_draft(draft)
            st.rerun()

    with col_restart:
        if st.button(
            "🔄 Reiniciar selección",
            use_container_width=True,
            key="mod_restart_v04",
        ):
            reset_draft()
            for key in [
                "mod_piece_v04",
                "mod_action_v04",
                "mod_destination_v04",
                "mod_entry_mode_v04",
                "mod_user_mode_v04",
                "mod_existing_user_v04",
                "mod_new_user_v04",
                "mod_include_twin_v08",
                "mod_doc_v08",
                "mod_range_v08",
            ]:
                st.session_state.pop(key, None)
            st.rerun()

    # --------------------------------------------------------
    # Propagación global
    # --------------------------------------------------------
    propagate = False
    occurrences = pd.DataFrame()

    if draft.get("replaced_from") and draft.get("replaced_to"):
        question(
            7,
            "¿Quieres reemplazar también en otros CECO?",
            (
                "Utiliza esta opción si la persona salió de la empresa "
                "y debe ser reemplazada en todas sus apariciones."
            ),
        )

        occurrences = occurrences_of_person(
            flow,
            draft["replaced_from"],
        )

        affected_cecos = (
            occurrences["CECO"].nunique()
            if not occurrences.empty
            else 0
        )
        affected_rows = (
            occurrences["_ID_FILA"].nunique()
            if not occurrences.empty
            else 0
        )
        affected_positions = len(occurrences)

        metric_cols = st.columns(3)
        metrics = [
            ("CECO donde participa", affected_cecos),
            ("Filas afectadas", affected_rows),
            ("Apariciones", affected_positions),
        ]

        for column, (label, value) in zip(metric_cols, metrics):
            with column:
                st.markdown(
                    compact_html(
                        f"""
                        <div class="metric-card">
                            <div class="metric-label">{escape(label)}</div>
                            <div class="metric-value">{value}</div>
                        </div>
                        """
                    ),
                    unsafe_allow_html=True,
                )

        with st.expander(
            (
                "Ver todas las apariciones de "
                f"{display_user(draft['replaced_from'], data)}"
            ),
            expanded=False,
        ):
            visible_occurrences = occurrences.drop(
                columns=["_ID_FILA"],
                errors="ignore",
            )
            st.dataframe(
                style_flow_table(
                    visible_occurrences.rename(
                        columns={
                            "Campo": "Lib1",
                            "ValorAntes": "Lib2",
                        }
                    ).assign(
                        Lib3="",
                        Lib4="",
                        Lib5="",
                    )
                )
                if not visible_occurrences.empty
                else visible_occurrences,
                use_container_width=True,
                hide_index=True,
            )

        propagate = st.radio(
            "¿Otros CECO?",
            options=["no", "yes"],
            format_func=lambda value: (
                "NO — guardar solo este tramo"
                if value == "no"
                else (
                    "SÍ — reemplazar en todos los CECO donde aparecía "
                    "la persona que salió"
                )
            ),
            key="mod_propagate_v04",
        ) == "yes"

    # --------------------------------------------------------
    # Guardado
    # --------------------------------------------------------
    question(
        8,
        "¿Guardar los cambios?",
        (
            "Los cambios se aplicarán a la base activa y quedarán "
            "registrados en la hoja Cambios."
        ),
    )

    target_row_ids = [
        int(value)
        for value in draft.get("target_row_ids", [])
    ]
    target_current_rows = flow[
        flow["_ID_FILA"].isin(target_row_ids)
    ]
    after_padded_for_check = libs_padded(libs_after)
    no_changes = bool(target_current_rows.empty) or all(
        libs_padded(libs_from_row(row)) == after_padded_for_check
        for _, row in target_current_rows.iterrows()
    )
    validation_errors = validate_flow_result(libs_after)
    missing_identification = (
        not clean_text(actor)
        or not clean_text(reason)
    )

    if validation_errors:
        for validation_error in validation_errors:
            st.error(validation_error)

    if missing_identification:
        st.warning(
            "Para guardar debes indicar quién realiza el cambio "
            "y el motivo."
        )

    save_clicked = st.button(
        "💾 Guardar cambios y preparar ZIP",
        type="primary",
        use_container_width=True,
        disabled=(
            no_changes
            or bool(validation_errors)
            or missing_identification
        ),
        key="mod_save_v06",
    )

    if no_changes:
        st.caption("Todavía no hay cambios para guardar.")

    if save_clicked:
        updated = flow.copy(deep=True)
        target_row_ids = [
            int(value)
            for value in draft.get("target_row_ids", [])
        ]
        selected_mask = updated["_ID_FILA"].isin(target_row_ids)

        if not selected_mask.any():
            st.error("Las filas seleccionadas ya no existen.")
            return

        after_padded = libs_padded(libs_after)
        timestamp = datetime.now(CHILE_TZ).strftime("%Y-%m-%d %H:%M:%S")
        changes: list[dict[str, Any]] = []

        for row_index, row in updated[selected_mask].iterrows():
            row_before = libs_padded(libs_from_row(row))

            for column, old_value, new_value in zip(
                LIB_COLS,
                row_before,
                after_padded,
            ):
                if strip_user(old_value) == strip_user(new_value):
                    continue

                updated.at[row_index, column] = strip_user(new_value)
                changes.append(
                    {
                        "FechaHora": timestamp,
                        "Usuario": actor or "anonimo",
                        "CECO": row["CECO"],
                        "Desde": row["Desde"],
                        "Hasta": row["Hasta"],
                        "TipoDoc": row["TipoDoc"],
                        "Campo": column,
                        "ValorAntes": strip_user(old_value) or "—",
                        "ValorDespues": strip_user(new_value) or "—",
                        "Nota": (
                            (reason or "Edición guiada de liberadores")
                            + (
                                " | aplicado también a CECO gemelo"
                                if len(set(draft.get("target_cecos", []))) > 1
                                else ""
                            )
                            + (
                                " | aplicado a ambos tipos de documento"
                                if len(set(draft.get("target_docs", []))) > 1
                                else ""
                            )
                        ),
                    }
                )


        if (
            propagate
            and draft.get("replaced_from")
            and draft.get("replaced_to")
        ):
            old_key = email_key(draft["replaced_from"])
            new_person = strip_user(draft["replaced_to"])

            for row_index, row in updated.iterrows():
                for column in LIB_COLS:
                    current = strip_user(row[column])

                    if email_key(current) != old_key:
                        continue

                    # Evita duplicar el registro ya modificado del tramo actual.
                    if int(row["_ID_FILA"]) in target_row_ids:
                        continue

                    updated.at[row_index, column] = new_person
                    changes.append(
                        {
                            "FechaHora": timestamp,
                            "Usuario": actor or "anonimo",
                            "CECO": row["CECO"],
                            "Desde": row["Desde"],
                            "Hasta": row["Hasta"],
                            "TipoDoc": row["TipoDoc"],
                            "Campo": column,
                            "ValorAntes": current,
                            "ValorDespues": new_person,
                            "Nota": (
                                (reason or "Reemplazo global")
                                + " | propagado a otros CECO"
                            ),
                        }
                    )

        if (
            propagate
            and draft.get("replaced_from")
            and draft.get("replaced_to")
        ):
            current_data = st.session_state.get(
                SESSION_DATA_KEY,
                data,
            )
            if isinstance(current_data, dict):
                updated_data, special_changes = (
                    replace_user_in_special_rules(
                        current_data,
                        draft["replaced_from"],
                        draft["replaced_to"],
                        actor,
                        reason,
                        timestamp,
                    )
                )
                st.session_state[SESSION_DATA_KEY] = updated_data
                changes.extend(special_changes)

        set_working_flow(updated)

        history = list(st.session_state.get(SESSION_HISTORY_KEY, []))
        history.extend(changes)
        st.session_state[SESSION_HISTORY_KEY] = history

        try:
            refresh_download(file_name, file_bytes)
            draft["libs_before"] = list(libs_after)
            draft["replaced_from"] = ""
            draft["replaced_to"] = ""
            draft["last_message"] = (
                f"Guardado correctamente: {len(changes)} cambio(s)."
            )
            set_draft(draft)

            st.success(
                f"Guardado correctamente: **{len(changes)} cambio(s)**."
            )
            st.toast("Base activa actualizada.", icon="✅")
        except ValueError as error:
            st.error(str(error))

    # --------------------------------------------------------
    # Descarga
    # --------------------------------------------------------
    render_download_selector("mod")

    history = st.session_state.get(SESSION_HISTORY_KEY, [])
    if history:
        with st.expander(
            f"Historial de esta sesión ({len(history)} cambios)",
            expanded=False,
        ):
            st.dataframe(
                pd.DataFrame(history),
                use_container_width=True,
                hide_index=True,
            )

    with st.expander("Restaurar versión original", expanded=False):
        st.warning(
            "Esta acción elimina las modificaciones realizadas "
            "durante la sesión."
        )

        confirm_restore = st.checkbox(
            "Confirmo que deseo restaurar la versión original cargada.",
            key="mod_confirm_restore_v04",
        )

        if st.button(
            "Restaurar versión original",
            disabled=not confirm_restore,
            use_container_width=True,
            key="mod_restore_v04",
        ):
            backup = st.session_state.get(SESSION_BACKUP_KEY)
            if isinstance(backup, pd.DataFrame):
                set_working_flow(backup.copy(deep=True))
                reset_draft()
                st.session_state[SESSION_HISTORY_KEY] = []
                st.session_state.pop(SESSION_DOWNLOAD_KEY, None)
                st.session_state.pop(SESSION_DOWNLOAD_NAME_KEY, None)
                st.session_state.pop(
                    SESSION_DOWNLOAD_PARQUET_KEY,
                    None,
                )
                st.session_state.pop(
                    SESSION_DOWNLOAD_CSV_KEY,
                    None,
                )
                st.session_state.pop(
                    SESSION_DOWNLOAD_PARQUET_NAME_KEY,
                    None,
                )
                st.session_state.pop(
                    SESSION_DOWNLOAD_CSV_NAME_KEY,
                    None,
                )
                st.session_state.pop(
                    SESSION_DOWNLOAD_EXCEL_KEY,
                    None,
                )
                st.session_state.pop(
                    SESSION_DOWNLOAD_EXCEL_NAME_KEY,
                    None,
                )
                st.rerun()


# ============================================================
# SIN ARCHIVO
# ============================================================

def render_no_file() -> None:
    st.warning(
        "No hay una versión activa. Primero carga los siete archivos desde "
        "**01 Cargar Versión**."
    )

    try:
        if st.button(
            "📤 Ir a 01 Cargar Versión",
            type="primary",
            use_container_width=True,
        ):
            st.switch_page("01_CARGAR_ARCHIVO_FLUJO.py")
    except Exception:
        st.info("Selecciona **01 Cargar Versión** desde la barra lateral.")


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    aplicar_estilos()
    render_header()

    data = st.session_state.get(SESSION_DATA_KEY)
    file_name = st.session_state.get(SESSION_FILE_KEY, {})
    file_bytes = st.session_state.get(
        SESSION_SOURCE_FILES_KEY,
        st.session_state.get(SESSION_FILE_BYTES_KEY, {}),
    )

    if (
        not isinstance(data, dict)
        or "flujo" not in data
        or not isinstance(data["flujo"], pd.DataFrame)
        or data["flujo"].empty
    ):
        render_no_file()
        return

    initialize_state(
        data=data,
        file_name=file_name,
        file_bytes=file_bytes,
    )

    render_wizard(
        data=data,
        file_name=file_name,
        file_bytes=file_bytes,
    )


if __name__ == "__main__":
    main()
