import re
import base64
import unicodedata
import math
from io import StringIO, BytesIO
from pathlib import Path
from urllib.parse import urljoin, unquote

import pandas as pd
import requests
import streamlit as st


# =========================
# Rutas del proyecto
# =========================

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
LOGO_PATH = PROJECT_DIR / "assets" / "logo.svg"


# =========================
# Configuración general
# =========================

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

MESES_NUM = {
    "Ene": 1,
    "Feb": 2,
    "Mar": 3,
    "Abr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Ago": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dic": 12,
    "Enero": 1,
    "Febrero": 2,
    "Marzo": 3,
    "Abril": 4,
    "Mayo": 5,
    "Junio": 6,
    "Julio": 7,
    "Agosto": 8,
    "Septiembre": 9,
    "Setiembre": 9,
    "Octubre": 10,
    "Noviembre": 11,
    "Diciembre": 12,
}

MESES_NOMBRE = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre"
}

MESES_ABREV = [
    "Ene", "Feb", "Mar", "Abr", "May", "Jun",
    "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"
]


# =========================
# Logo
# =========================

def mostrar_logo_centrado():
    if LOGO_PATH.exists():
        logo_svg = LOGO_PATH.read_text(encoding="utf-8")
        logo_base64 = base64.b64encode(
            logo_svg.encode("utf-8")
        ).decode("utf-8")

        st.markdown(
            f"""
            <div style="
                width: 100%;
                display: flex;
                justify-content: center;
                align-items: center;
                margin-top: 10px;
                margin-bottom: 20px;
            ">
                <img
                    src="data:image/svg+xml;base64,{logo_base64}"
                    style="width: 260px; display: block;"
                >
            </div>
            """,
            unsafe_allow_html=True
        )
    else:
        st.warning(f"Logo no encontrado: {LOGO_PATH}")


# =========================
# Utilidades
# =========================

def normalizar_texto(texto):
    if pd.isna(texto):
        return ""

    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto)

    return texto


def convertir_numerico(serie):
    return pd.to_numeric(
        serie,
        errors="coerce"
    )


def descargar_url_excel(url):
    response = requests.get(
        url,
        headers=HEADERS,
        timeout=30
    )

    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "").lower()

    es_excel = (
        "spreadsheet" in content_type
        or "excel" in content_type
        or ".xlsx" in url.lower()
        or ".xls" in url.lower()
    )

    if not es_excel:
        raise ValueError(
            f"La URL no parece ser Excel. Content-Type: {content_type}"
        )

    return response.content


def buscar_links_excel_en_pagina(url_pagina, patrones):
    response = requests.get(
        url_pagina,
        headers=HEADERS,
        timeout=30
    )

    response.raise_for_status()

    html = response.text

    hrefs = re.findall(
        r'href=["\'](.*?)["\']',
        html,
        flags=re.IGNORECASE
    )

    enlaces = []

    for href in hrefs:
        url_completa = urljoin(url_pagina, href)
        url_decodificada = unquote(url_completa)
        url_norm = normalizar_texto(url_decodificada)

        if ".xls" not in url_norm:
            continue

        if any(normalizar_texto(patron) in url_norm for patron in patrones):
            enlaces.append(url_completa)

    enlaces = list(dict.fromkeys(enlaces))

    return enlaces


def encontrar_hoja_por_columnas(contenido_excel, columnas_objetivo):
    excel = pd.ExcelFile(BytesIO(contenido_excel))

    columnas_objetivo_norm = [
        normalizar_texto(col)
        for col in columnas_objetivo
    ]

    for hoja in excel.sheet_names:
        try:
            df_preview = pd.read_excel(
                BytesIO(contenido_excel),
                sheet_name=hoja,
                nrows=20
            )

            columnas_preview_norm = [
                normalizar_texto(col)
                for col in df_preview.columns
            ]

            if all(col in columnas_preview_norm for col in columnas_objetivo_norm):
                return hoja

        except Exception:
            continue

    if "General" in excel.sheet_names:
        return "General"

    return excel.sheet_names[0]


# =========================
# DÓLAR SII
# =========================

def encontrar_tabla_dolar(tablas):
    meses_abreviados = set(MESES_ABREV)

    for tabla in tablas:
        columnas = [str(col).strip() for col in tabla.columns]

        tiene_columna_dia = "Día" in columnas or "Dia" in columnas
        meses_en_columnas = meses_abreviados.intersection(columnas)
        texto_tabla = tabla.astype(str).to_string()
        tiene_promedio = "Promedio" in texto_tabla

        if tiene_columna_dia and len(meses_en_columnas) >= 8 and tiene_promedio:
            return tabla

    return None


@st.cache_data
def cargar_tabla_dolar(anio):
    url = f"https://www.sii.cl/valores_y_fechas/dolar/dolar{anio}.htm"

    response = requests.get(
        url,
        headers=HEADERS,
        timeout=30
    )
    response.raise_for_status()

    tablas = pd.read_html(
        StringIO(response.text),
        decimal=",",
        thousands="."
    )

    return encontrar_tabla_dolar(tablas)


def obtener_resumen_mes_dolar(tabla, mes):
    tabla_temp = tabla.copy()

    columna_dia = "Día" if "Día" in tabla_temp.columns else "Dia"

    fila_promedio = tabla_temp[
        tabla_temp[columna_dia].astype(str).str.strip().str.lower() == "promedio"
    ]

    if not fila_promedio.empty:
        valor_promedio = fila_promedio[mes].iloc[0]
    else:
        valor_promedio = tabla_temp[mes].mean(skipna=True)

    tabla_dias = tabla_temp[
        pd.to_numeric(tabla_temp[columna_dia], errors="coerce").notna()
    ].copy()

    valores_mes = tabla_dias[[columna_dia, mes]].dropna(subset=[mes])

    if valores_mes.empty:
        ultimo_dia = None
        ultimo_valor = None
    else:
        ultimo_dia = valores_mes[columna_dia].iloc[-1]
        ultimo_valor = valores_mes[mes].iloc[-1]

    return ultimo_dia, ultimo_valor, valor_promedio


def generar_df_dolar(anios):
    registros = []

    for anio in anios:
        try:
            tabla_dolar = cargar_tabla_dolar(anio)

            if tabla_dolar is None:
                continue

            for mes in MESES_ABREV:
                if mes not in tabla_dolar.columns:
                    continue

                ultimo_dia, ultimo_valor, valor_promedio = obtener_resumen_mes_dolar(
                    tabla_dolar,
                    mes
                )

                mes_numero = MESES_NUM[mes]

                registros.append({
                    "Año": anio,
                    "Mes": mes_numero,
                    "Fecha": pd.Timestamp(year=anio, month=mes_numero, day=1),
                    "Dolar promedio SII": valor_promedio,
                    "Dolar ultimo observado SII": ultimo_valor,
                    "Dolar ultimo dia observado": ultimo_dia
                })

        except Exception as e:
            st.warning(f"Dólar SII {anio}: {e}")

    df = pd.DataFrame(registros)

    if not df.empty:
        df["Dolar promedio SII"] = convertir_numerico(df["Dolar promedio SII"])
        df["Dolar ultimo observado SII"] = convertir_numerico(
            df["Dolar ultimo observado SII"]
        )

    return df


# =========================
# UTM SII
# =========================

def encontrar_tabla_utm(tablas):
    for tabla in tablas:
        texto = normalizar_texto(tabla.astype(str).to_string())

        tiene_mes = "mes" in texto or "enero" in texto
        tiene_utm = "utm" in texto
        tiene_uta = "uta" in texto
        tiene_ipc = "ipc" in texto

        if tiene_mes and tiene_utm and tiene_uta and tiene_ipc:
            return tabla.copy()

    return None


@st.cache_data
def cargar_tabla_utm(anio):
    url = f"https://www.sii.cl/valores_y_fechas/utm/utm{anio}.htm"

    response = requests.get(
        url,
        headers=HEADERS,
        timeout=30
    )
    response.raise_for_status()

    tablas = pd.read_html(
        StringIO(response.text),
        decimal=",",
        thousands="."
    )

    tabla_utm = encontrar_tabla_utm(tablas)

    if tabla_utm is None:
        raise ValueError("No se encontró una tabla UTM compatible.")

    if len(tabla_utm.columns) < 7:
        raise ValueError("La tabla UTM tiene menos columnas de las esperadas.")

    tabla_utm = tabla_utm.iloc[:, :7].copy()

    tabla_utm.columns = [
        "Mes texto",
        "UTM",
        "UTA",
        "IPC valor puntos SII",
        "UTM variacion mensual",
        "UTM variacion acumulado",
        "UTM variacion 12 meses"
    ]

    tabla_utm.insert(0, "Año", anio)

    return tabla_utm


def generar_df_utm(anios):
    tablas = []

    for anio in anios:
        try:
            tabla = cargar_tabla_utm(anio)
            tablas.append(tabla)

        except Exception as e:
            st.warning(f"UTM SII {anio}: {e}")

    if not tablas:
        return pd.DataFrame()

    df = pd.concat(tablas, ignore_index=True)

    df["Mes"] = df["Mes texto"].astype(str).str.strip().map(MESES_NUM)

    df = df.dropna(subset=["Año", "Mes"]).copy()

    df["Año"] = df["Año"].astype(int)
    df["Mes"] = df["Mes"].astype(int)

    df["Fecha"] = pd.to_datetime(
        df["Año"].astype(str)
        + "-"
        + df["Mes"].astype(str)
        + "-01"
    )

    columnas_numericas = [
        "UTM",
        "UTA",
        "IPC valor puntos SII",
        "UTM variacion mensual",
        "UTM variacion acumulado",
        "UTM variacion 12 meses"
    ]

    for columna in columnas_numericas:
        if columna in df.columns:
            df[columna] = convertir_numerico(df[columna])

    columnas_salida = [
        "Fecha",
        "Año",
        "Mes",
        "UTM",
        "UTA",
        "IPC valor puntos SII",
        "UTM variacion mensual",
        "UTM variacion acumulado",
        "UTM variacion 12 meses"
    ]

    return df[columnas_salida]


# =========================
# IPC INE
# =========================

URL_PAGINA_INE_IPC = (
    "https://www.ine.gob.cl/estadisticas-por-tema/"
    "precios-e-inflacion/indice-de-precios-al-consumidor"
)

URL_EXCEL_IPC_DIRECTA = (
    "https://www.ine.gob.cl/docs/default-source/"
    "%C3%ADndice-de-precios-al-consumidor/cuadros-estadisticos/"
    "base-anual-2023_100/series-de-tiempo/"
    "ipc-xls.xlsx?sfvrsn=5b901f39_70"
)

URL_EXCEL_IPC_BASE = (
    "https://www.ine.gob.cl/docs/default-source/"
    "%C3%ADndice-de-precios-al-consumidor/cuadros-estadisticos/"
    "base-anual-2023_100/series-de-tiempo/"
    "ipc-xls.xlsx"
)


@st.cache_data
def descargar_excel_ipc_ine():
    errores = []

    for url in [URL_EXCEL_IPC_DIRECTA, URL_EXCEL_IPC_BASE]:
        try:
            return descargar_url_excel(url)
        except Exception as e:
            errores.append(f"{url}: {e}")

    try:
        enlaces = buscar_links_excel_en_pagina(
            URL_PAGINA_INE_IPC,
            patrones=[
                "ipc-xls",
                "series-de-tiempo",
                "ipc"
            ]
        )

        for enlace in enlaces:
            try:
                return descargar_url_excel(enlace)
            except Exception as e:
                errores.append(f"{enlace}: {e}")

    except Exception as e:
        errores.append(f"Búsqueda automática IPC: {e}")

    raise ValueError(
        "No se pudo descargar el Excel IPC INE. "
        + " | ".join(errores[:5])
    )


def encontrar_hoja_ipc(contenido_excel):
    excel = pd.ExcelFile(BytesIO(contenido_excel))

    for hoja in excel.sheet_names:
        try:
            df_raw = pd.read_excel(
                BytesIO(contenido_excel),
                sheet_name=hoja,
                header=None,
                nrows=30
            )

            contiene_header = df_raw.apply(
                lambda fila: (
                    fila.astype(str).str.strip().eq("Año").any()
                    and fila.astype(str).str.strip().eq("Mes").any()
                    and fila.astype(str).str.strip().eq("Glosa").any()
                ),
                axis=1
            ).any()

            if contiene_header:
                return hoja

        except Exception:
            continue

    return excel.sheet_names[0]


def leer_excel_ipc_limpio(contenido_excel):
    hoja_ipc = encontrar_hoja_ipc(contenido_excel)

    df_raw = pd.read_excel(
        BytesIO(contenido_excel),
        sheet_name=hoja_ipc,
        header=None
    )

    filas_header = df_raw[
        df_raw.apply(
            lambda fila: (
                fila.astype(str).str.strip().eq("Año").any()
                and fila.astype(str).str.strip().eq("Mes").any()
                and fila.astype(str).str.strip().eq("Glosa").any()
            ),
            axis=1
        )
    ]

    if filas_header.empty:
        raise ValueError("No se encontró fila de encabezado IPC con Año, Mes y Glosa.")

    fila_header = filas_header.index[0]

    columnas = df_raw.iloc[fila_header].tolist()

    df_ipc = df_raw.iloc[fila_header + 1:].copy()
    df_ipc.columns = columnas
    df_ipc = df_ipc.dropna(how="all").reset_index(drop=True)

    df_ipc.columns = [
        str(col).strip()
        for col in df_ipc.columns
    ]

    return df_ipc


def generar_df_ipc_ine():
    try:
        contenido = descargar_excel_ipc_ine()
        df_ipc = leer_excel_ipc_limpio(contenido)

        df = df_ipc[
            df_ipc["Glosa"].astype(str).str.strip().eq("IPC General")
        ].copy()

        columnas_numericas = [
            "Año",
            "Mes",
            "Índice",
            "Variación Mensual (%)",
            "Variación Acumulada (%)",
            "Variación 12 Meses (%)"
        ]

        for columna in columnas_numericas:
            if columna in df.columns:
                df[columna] = convertir_numerico(df[columna])

        df = df.dropna(subset=["Año", "Mes", "Índice"]).copy()

        df["Año"] = df["Año"].astype(int)
        df["Mes"] = df["Mes"].astype(int)

        df["Fecha"] = pd.to_datetime(
            df["Año"].astype(str)
            + "-"
            + df["Mes"].astype(str)
            + "-01"
        )

        df = df.rename(
            columns={
                "Índice": "IPC INE indice",
                "Variación Mensual (%)": "IPC INE variacion mensual",
                "Variación Acumulada (%)": "IPC INE variacion acumulada",
                "Variación 12 Meses (%)": "IPC INE variacion 12 meses"
            }
        )

        columnas_salida = [
            "Fecha",
            "Año",
            "Mes",
            "IPC INE indice",
            "IPC INE variacion mensual",
            "IPC INE variacion acumulada",
            "IPC INE variacion 12 meses"
        ]

        return df[columnas_salida].sort_values("Fecha")

    except Exception as e:
        st.warning(f"IPC INE: {e}")
        return pd.DataFrame()


# =========================
# ICL / IR INE
# =========================

URL_PAGINA_INE_REMUNERACIONES = (
    "https://www.ine.gob.cl/estadisticas-por-tema/"
    "mercado-laboral/remuneraciones-y-costos-laborales"
)

URL_ARCHIVO_ICL_DIRECTO = (
    "https://www.ine.gob.cl/docs/default-source/sueldos-y-salarios/"
    "cuadros-estadisticos/ir-icl-base-anual-2023-100/"
    "series-base-2023/tabulado_icl.xlsx?sfvrsn=43d76e7c_50"
)

URL_ARCHIVO_ICL_BASE = (
    "https://www.ine.gob.cl/docs/default-source/sueldos-y-salarios/"
    "cuadros-estadisticos/ir-icl-base-anual-2023-100/"
    "series-base-2023/tabulado_icl.xlsx"
)

URL_ARCHIVO_IR_DIRECTO = (
    "https://www.ine.gob.cl/docs/default-source/sueldos-y-salarios/"
    "cuadros-estadisticos/ir-icl-base-anual-2023-100/"
    "series-base-2023/tabulado_ir.xlsx?sfvrsn=77d2fe83_52"
)

URL_ARCHIVO_IR_BASE = (
    "https://www.ine.gob.cl/docs/default-source/sueldos-y-salarios/"
    "cuadros-estadisticos/ir-icl-base-anual-2023-100/"
    "series-base-2023/tabulado_ir.xlsx"
)


@st.cache_data
def descargar_excel_icl():
    errores = []

    for url in [URL_ARCHIVO_ICL_DIRECTO, URL_ARCHIVO_ICL_BASE]:
        try:
            return descargar_url_excel(url)
        except Exception as e:
            errores.append(f"{url}: {e}")

    try:
        enlaces = buscar_links_excel_en_pagina(
            URL_PAGINA_INE_REMUNERACIONES,
            patrones=[
                "tabulado_icl",
                "icl",
                "ir-icl",
                "series-base"
            ]
        )

        for enlace in enlaces:
            try:
                return descargar_url_excel(enlace)
            except Exception as e:
                errores.append(f"{enlace}: {e}")

    except Exception as e:
        errores.append(f"Búsqueda automática ICL: {e}")

    raise ValueError(
        "No se pudo descargar el Excel ICL INE. "
        + " | ".join(errores[:5])
    )


@st.cache_data
def descargar_excel_ir():
    errores = []

    for url in [URL_ARCHIVO_IR_DIRECTO, URL_ARCHIVO_IR_BASE]:
        try:
            return descargar_url_excel(url)
        except Exception as e:
            errores.append(f"{url}: {e}")

    try:
        enlaces = buscar_links_excel_en_pagina(
            URL_PAGINA_INE_REMUNERACIONES,
            patrones=[
                "tabulado_ir",
                "tabulado ir",
                "ir-icl",
                "series-base",
                "remuneraciones"
            ]
        )

        enlaces_priorizados = sorted(
            enlaces,
            key=lambda x: (
                "tabulado_ir" not in normalizar_texto(unquote(x)),
                "series-base" not in normalizar_texto(unquote(x))
            )
        )

        for enlace in enlaces_priorizados:
            try:
                return descargar_url_excel(enlace)
            except Exception as e:
                errores.append(f"{enlace}: {e}")

    except Exception as e:
        errores.append(f"Búsqueda automática IR: {e}")

    raise ValueError(
        "No se pudo descargar el Excel IR INE. "
        + " | ".join(errores[:5])
    )


def generar_df_indice_remuneraciones(
    contenido,
    nombre_columna_indice,
    prefijo_salida
):
    hoja = encontrar_hoja_por_columnas(
        contenido,
        columnas_objetivo=["año", "mes", "índice"]
    )

    df = pd.read_excel(
        BytesIO(contenido),
        sheet_name=hoja
    )

    columnas_actuales = {
        normalizar_texto(col): col
        for col in df.columns
    }

    columnas_necesarias = ["ano", "mes", "indice"]

    for columna in columnas_necesarias:
        if columna not in columnas_actuales:
            st.warning(
                f"{prefijo_salida}: no se encontró columna requerida '{columna}'. "
                f"Columnas disponibles: {list(columnas_actuales.keys())}"
            )
            return pd.DataFrame()

    df = df.rename(
        columns={
            columnas_actuales["ano"]: "Año",
            columnas_actuales["mes"]: "Mes",
            columnas_actuales["indice"]: nombre_columna_indice
        }
    )

    df["Año"] = convertir_numerico(df["Año"])
    df["Mes"] = convertir_numerico(df["Mes"])
    df[nombre_columna_indice] = convertir_numerico(df[nombre_columna_indice])

    df = df.dropna(
        subset=["Año", "Mes", nombre_columna_indice]
    ).copy()

    df["Año"] = df["Año"].astype(int)
    df["Mes"] = df["Mes"].astype(int)

    df["Fecha"] = pd.to_datetime(
        df["Año"].astype(str)
        + "-"
        + df["Mes"].astype(str)
        + "-01"
    )

    columnas_extra = {}

    for original, nuevo in {
        "var_mensual": f"{prefijo_salida} variacion mensual",
        "var_acum": f"{prefijo_salida} variacion acumulada",
        "var_12": f"{prefijo_salida} variacion 12 meses"
    }.items():
        if original in columnas_actuales:
            df[nuevo] = convertir_numerico(
                df[columnas_actuales[original]]
            )
            columnas_extra[nuevo] = nuevo

    columnas_salida = [
        "Fecha",
        "Año",
        "Mes",
        nombre_columna_indice
    ] + list(columnas_extra.keys())

    return df[columnas_salida].sort_values("Fecha")


def generar_df_icl():
    try:
        contenido = descargar_excel_icl()

        return generar_df_indice_remuneraciones(
            contenido=contenido,
            nombre_columna_indice="ICL INE indice",
            prefijo_salida="ICL INE"
        )

    except Exception as e:
        st.warning(f"ICL INE: {e}")
        return pd.DataFrame()


def generar_df_ir():
    try:
        contenido = descargar_excel_ir()

        return generar_df_indice_remuneraciones(
            contenido=contenido,
            nombre_columna_indice="IR INE indice",
            prefijo_salida="IR INE"
        )

    except Exception as e:
        st.warning(f"IR INE: {e}")
        return pd.DataFrame()


# =========================
# MOP
# =========================

URL_PAGINA_MOP = (
    "https://planeamiento.mop.gob.cl/"
    "indices-y-precios-para-calculo-del-reajuste-polinomico/"
)

MESES_MOP = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12
}

CONCEPTOS_OBJETIVO_MOP = {
    "MOP indice precios consumidor": [
        "Índice de precios al consumidor (1)",
        "Índice de precios al consumidor",
        "IPC"
    ],
    "MOP indice remuneraciones": [
        "Índice de remuneraciones (2)",
        "Índice de remuneraciones",
        "Remuneraciones"
    ],
    "MOP petroleo diesel": [
        "Petróleo Diesel (4)",
        "Petróleo Diesel",
        "Petroleo Diesel"
    ],
    "MOP dolar observado": [
        "Dólar observado",
        "Dolar observado"
    ],
    "MOP petroleo diesel refineria concon": [
        "Petróleo Diesel de refinería CONCÓN",
        "Petroleo Diesel de refineria CONCON",
        "Diesel de refinería CONCÓN",
        "Diesel refinería CONCÓN"
    ]
}


def extraer_mes_anio_desde_archivo_mop(archivo):
    nombre = normalizar_texto(archivo)

    nombre_limpio = re.sub(r"[_\-.]+", " ", nombre)
    nombre_limpio = re.sub(r"\s+", " ", nombre_limpio).strip()

    partes = nombre_limpio.split()

    mes_detectado = None
    numero_mes = None

    for mes, numero in MESES_MOP.items():
        if mes in partes:
            mes_detectado = mes
            numero_mes = numero
            break

    if numero_mes is None:
        match_mes_num = re.search(
            r"(?:^|[^0-9])(0?[1-9]|1[0-2])(?:[^0-9]|$)",
            nombre_limpio
        )

        if match_mes_num:
            numero_mes = int(match_mes_num.group(1))
            mes_detectado = MESES_NOMBRE.get(numero_mes, "")

    anio_match = re.search(r"(20\d{2})", nombre_limpio)
    anio_detectado = int(anio_match.group(1)) if anio_match else None

    return mes_detectado, numero_mes, anio_detectado


def obtener_valor_conceptos_mop(df, variantes_concepto):
    variantes_norm = [
        normalizar_texto(concepto)
        for concepto in variantes_concepto
    ]

    for _, fila in df.iterrows():
        detalle = normalizar_texto(fila[1]) if len(fila) > 1 else ""

        for concepto_norm in variantes_norm:
            if concepto_norm and concepto_norm in detalle:
                if len(fila) > 3:
                    return fila[3]
                return pd.NA

    return pd.NA


@st.cache_data
def obtener_archivos_excel_mop():
    response = requests.get(
        URL_PAGINA_MOP,
        headers=HEADERS,
        timeout=30
    )
    response.raise_for_status()

    html = response.text

    hrefs = re.findall(
        r'href=["\'](.*?)["\']',
        html,
        flags=re.IGNORECASE
    )

    enlaces = []

    for href in hrefs:
        url_completa = urljoin(URL_PAGINA_MOP, href)
        url_decodificada = unquote(url_completa)

        enlaces.append({
            "url": url_completa,
            "url_decodificada": url_decodificada
        })

    df_enlaces = pd.DataFrame(enlaces).drop_duplicates()

    df_excel = df_enlaces[
        df_enlaces["url_decodificada"]
        .str.lower()
        .str.contains(r"\.xls|\.xlsx", regex=True, na=False)
    ].copy()

    if df_excel.empty:
        return df_excel

    df_excel["archivo"] = df_excel["url_decodificada"].str.split("/").str[-1]

    df_excel[["mes_texto", "Mes", "Año"]] = df_excel["archivo"].apply(
        lambda x: pd.Series(extraer_mes_anio_desde_archivo_mop(x))
    )

    df_excel = df_excel.dropna(subset=["Año", "Mes"]).copy()

    df_excel["Año"] = df_excel["Año"].astype(int)
    df_excel["Mes"] = df_excel["Mes"].astype(int)

    df_excel = df_excel.sort_values(["Año", "Mes"]).reset_index(drop=True)

    return df_excel


@st.cache_data
def leer_archivo_excel_mop(url_archivo, archivo):
    response = requests.get(
        url_archivo,
        headers=HEADERS,
        timeout=30
    )
    response.raise_for_status()

    contenido = response.content

    excel = pd.ExcelFile(BytesIO(contenido))

    hoja = "planilla" if "planilla" in excel.sheet_names else excel.sheet_names[0]

    df = pd.read_excel(
        BytesIO(contenido),
        sheet_name=hoja,
        header=None
    )

    mes_texto, mes_numero, anio = extraer_mes_anio_desde_archivo_mop(archivo)

    registro = {
        "Año": anio,
        "Mes": mes_numero,
        "Fecha": pd.Timestamp(year=anio, month=mes_numero, day=1)
    }

    for nombre_columna, variantes in CONCEPTOS_OBJETIVO_MOP.items():
        registro[nombre_columna] = obtener_valor_conceptos_mop(
            df,
            variantes
        )

    return registro


def generar_df_mop(anios):
    try:
        df_archivos = obtener_archivos_excel_mop()

        if df_archivos.empty:
            return pd.DataFrame()

        df_archivos = df_archivos[
            df_archivos["Año"].isin(anios)
        ].copy()

        if df_archivos.empty:
            return pd.DataFrame()

        registros = []

        barra = st.progress(0)
        estado = st.empty()

        total = len(df_archivos)

        for posicion, (_, row) in enumerate(df_archivos.iterrows(), start=1):
            estado.write(f"Procesando MOP {posicion} de {total}")

            try:
                registro = leer_archivo_excel_mop(
                    row["url_decodificada"],
                    row["archivo"]
                )
                registros.append(registro)
            except Exception as e:
                st.warning(f"MOP {row['archivo']}: {e}")

            barra.progress(posicion / total)

        barra.empty()
        estado.empty()

        df = pd.DataFrame(registros)

        if df.empty:
            return df

        for columna in df.columns:
            if columna not in ["Fecha", "Año", "Mes"]:
                df[columna] = convertir_numerico(df[columna])

        return df.sort_values("Fecha")

    except Exception as e:
        st.warning(f"MOP: {e}")
        return pd.DataFrame()


# =========================
# Consolidado
# =========================

def unir_dataframes_temporales(dataframes):
    dataframes_validos = [
        df for df in dataframes
        if df is not None and not df.empty
    ]

    if not dataframes_validos:
        return pd.DataFrame()

    df_final = dataframes_validos[0].copy()

    for df in dataframes_validos[1:]:
        df_final = pd.merge(
            df_final,
            df,
            on=["Fecha", "Año", "Mes"],
            how="outer"
        )

    df_final = df_final.sort_values(["Fecha"]).reset_index(drop=True)

    return df_final


def agregar_mes_nombre(df):
    df = df.copy()

    if "Mes" in df.columns:
        df["Mes nombre"] = df["Mes"].map(MESES_NOMBRE)

    return df


def preparar_consolidado_resumido(df_consolidado):
    columnas_resumidas = [
        "Año",
        "Mes",
        "Mes nombre",
        "Fecha",
        "Dolar promedio SII",
        "Dolar ultimo observado SII",
        "Dolar ultimo dia observado",
        "UTM",
        "IPC valor puntos SII",
        "IPC INE indice",
        "ICL INE indice",
        "IR INE indice",
        "MOP indice precios consumidor",
        "MOP indice remuneraciones",
        "MOP petroleo diesel",
        "MOP dolar observado",
        "MOP petroleo diesel refineria concon"
    ]

    df = df_consolidado.copy()
    df = agregar_mes_nombre(df)

    for columna in columnas_resumidas:
        if columna not in df.columns:
            df[columna] = pd.NA

    df = df[columnas_resumidas].copy()

    df = df.sort_values(
        ["Año", "Mes"],
        ascending=[True, True]
    ).reset_index(drop=True)

    return df


# =========================
# Cálculo automático Ítem 2 e Ítem 3
# =========================

def truncar_15_decimales(valor):
    if pd.isna(valor):
        return pd.NA

    factor = 10 ** 15

    return math.trunc(float(valor) * factor) / factor


def calcular_item_2_item_3(df_resumido, columna_indice_item_2):
    df = df_resumido.copy()

    df = df.sort_values(
        ["Año", "Mes"],
        ascending=[True, True]
    ).reset_index(drop=True)

    columnas_requeridas = [
        columna_indice_item_2,
        "UTM",
        "MOP petroleo diesel refineria concon"
    ]

    for columna in columnas_requeridas:
        if columna not in df.columns:
            df[columna] = pd.NA

    df[columna_indice_item_2] = pd.to_numeric(
        df[columna_indice_item_2],
        errors="coerce"
    )

    df["UTM"] = pd.to_numeric(
        df["UTM"],
        errors="coerce"
    )

    df["MOP petroleo diesel refineria concon"] = pd.to_numeric(
        df["MOP petroleo diesel refineria concon"],
        errors="coerce"
    )

    df["Item 2 calculado"] = pd.NA

    item_2_anterior = pd.NA

    for i in range(len(df)):
        if i < 3:
            continue

        indice_i_menos_2 = df.loc[i - 2, columna_indice_item_2]
        indice_i_menos_3 = df.loc[i - 3, columna_indice_item_2]

        if pd.isna(item_2_anterior):
            item_2_anterior = indice_i_menos_3

        if (
            pd.isna(indice_i_menos_2)
            or pd.isna(indice_i_menos_3)
            or pd.isna(item_2_anterior)
            or indice_i_menos_3 == 0
        ):
            df.loc[i, "Item 2 calculado"] = pd.NA
            continue

        factor_indice = truncar_15_decimales(
            indice_i_menos_2 / indice_i_menos_3
        )

        item_2_actual = factor_indice * item_2_anterior

        df.loc[i, "Item 2 calculado"] = item_2_actual

        item_2_anterior = item_2_actual

    df["Item 3 calculado"] = (
        df["MOP petroleo diesel refineria concon"] * 1.05327007171179
        + 1.5 * df["UTM"]
    )

    return df


def estilo_columnas_calculadas(df):
    columnas_calculadas = [
        "Item 2 calculado",
        "Item 3 calculado"
    ]

    def aplicar_estilo(columna):
        if columna.name in columnas_calculadas:
            return [
                "background-color: #FFF2CC; font-weight: bold;"
                for _ in columna
            ]

        return ["" for _ in columna]

    return df.style.apply(aplicar_estilo, axis=0)


# =========================
# Excels
# =========================

def crear_excel_consolidado(df_consolidado, df_resumido=None):
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_consolidado.to_excel(
            writer,
            sheet_name="Consolidado completo",
            index=False
        )

        if df_resumido is not None:
            df_resumido.to_excel(
                writer,
                sheet_name="Consolidado resumido",
                index=False
            )

    return buffer.getvalue()


def crear_excel_consolidado_resumido(df_resumido):
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_resumido.to_excel(
            writer,
            sheet_name="Consolidado resumido",
            index=False
        )

    return buffer.getvalue()


def crear_excel_consolidado_calculado(df_calculado):
    from openpyxl.styles import PatternFill, Font

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_calculado.to_excel(
            writer,
            sheet_name="Consolidado calculado",
            index=False
        )

        worksheet = writer.sheets["Consolidado calculado"]

        color_calculado = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid"
        )

        fuente_bold = Font(bold=True)

        columnas_calculadas = [
            "Item 2 calculado",
            "Item 3 calculado"
        ]

        for idx, celda in enumerate(worksheet[1], start=1):
            if celda.value in columnas_calculadas:
                celda.fill = color_calculado
                celda.font = fuente_bold

                for fila in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=fila, column=idx).fill = color_calculado

    return buffer.getvalue()


# =========================
# App Streamlit
# =========================

st.set_page_config(
    page_title="Consolidado Temporal ENAEX",
    page_icon="🏢",
    layout="wide"
)

mostrar_logo_centrado()

st.markdown(
    "<h1 style='text-align: center;'>Consolidado Temporal de Indicadores</h1>",
    unsafe_allow_html=True
)

st.markdown(
    """
    <p style='text-align: center; font-size: 18px;'>
        Genera una base unificada mensual con indicadores SII, INE y MOP.
    </p>
    """,
    unsafe_allow_html=True
)

st.markdown("---")


# =========================
# Parámetros
# =========================

anios_disponibles = list(range(2009, 2027))
ultimos_3_anios = anios_disponibles[-3:]

st.subheader("Selecciona los años")

columnas_checkbox = st.columns(6)

anios_seleccionados = []

for posicion, anio in enumerate(anios_disponibles):
    columna_actual = columnas_checkbox[posicion % 6]

    with columna_actual:
        seleccionado = st.checkbox(
            str(anio),
            value=(anio in ultimos_3_anios),
            key=f"checkbox_anio_consolidado_{anio}"
        )

        if seleccionado:
            anios_seleccionados.append(anio)

st.write("Años seleccionados:", anios_seleccionados)

with st.expander("Fuentes a incluir"):
    incluir_dolar = st.checkbox("Dólar SII", value=True)
    incluir_utm = st.checkbox("UTM SII / IPC valor puntos SII", value=True)
    incluir_ipc_ine = st.checkbox("IPC INE", value=True)
    incluir_icl = st.checkbox("ICL INE", value=True)
    incluir_ir = st.checkbox("IR INE", value=True)
    incluir_mop = st.checkbox("MOP reajuste polinómico", value=True)


# =========================
# Botón generar
# =========================

if st.button("Generar consolidado temporal"):
    if not anios_seleccionados:
        st.warning("Debes seleccionar al menos un año.")
    else:
        dataframes = []

        with st.spinner("Generando consolidado temporal..."):
            if incluir_dolar:
                st.write("Consultando Dólar SII...")
                df_dolar = generar_df_dolar(anios_seleccionados)
                dataframes.append(df_dolar)

            if incluir_utm:
                st.write("Consultando UTM SII...")
                df_utm = generar_df_utm(anios_seleccionados)
                dataframes.append(df_utm)

            if incluir_ipc_ine:
                st.write("Consultando IPC INE...")
                df_ipc_ine = generar_df_ipc_ine()

                if not df_ipc_ine.empty:
                    df_ipc_ine = df_ipc_ine[
                        df_ipc_ine["Año"].isin(anios_seleccionados)
                    ].copy()

                dataframes.append(df_ipc_ine)

            if incluir_icl:
                st.write("Consultando ICL INE...")
                df_icl = generar_df_icl()

                if not df_icl.empty:
                    df_icl = df_icl[
                        df_icl["Año"].isin(anios_seleccionados)
                    ].copy()

                dataframes.append(df_icl)

            if incluir_ir:
                st.write("Consultando IR INE...")
                df_ir = generar_df_ir()

                if not df_ir.empty:
                    df_ir = df_ir[
                        df_ir["Año"].isin(anios_seleccionados)
                    ].copy()

                dataframes.append(df_ir)

            if incluir_mop:
                st.write("Consultando MOP...")
                df_mop = generar_df_mop(anios_seleccionados)
                dataframes.append(df_mop)

            df_consolidado = unir_dataframes_temporales(dataframes)

        st.session_state["df_consolidado_temporal"] = df_consolidado

        if df_consolidado.empty:
            st.warning("No se generaron datos para el consolidado.")
        else:
            st.success("Consolidado temporal generado correctamente.")


# =========================
# Mostrar resultados
# =========================

if "df_consolidado_temporal" in st.session_state:
    df_consolidado = st.session_state["df_consolidado_temporal"]

    if not df_consolidado.empty:
        st.markdown("---")

        st.subheader("Tabla consolidada completa")

        st.dataframe(
            df_consolidado,
            use_container_width=True
        )

        df_consolidado_resumido = preparar_consolidado_resumido(df_consolidado)

        st.subheader("Consolidado resumido")

        st.dataframe(
            df_consolidado_resumido,
            use_container_width=True
        )

        st.markdown("---")

        calcular_items = st.checkbox(
            "Calcular automáticamente Ítem 2 e Ítem 3 usando el consolidado resumido",
            value=False
        )

        df_consolidado_calculado = None

        if calcular_items:
            st.info(
                "Ítem 2 se calcula automáticamente usando el índice seleccionado. "
                "Ítem 3 se calcula usando MOP petróleo diesel refinería Concón y UTM."
            )

            opciones_indice_item_2 = []

            if "IR INE indice" in df_consolidado_resumido.columns:
                if df_consolidado_resumido["IR INE indice"].notna().any():
                    opciones_indice_item_2.append("IR INE indice")

            if "ICL INE indice" in df_consolidado_resumido.columns:
                if df_consolidado_resumido["ICL INE indice"].notna().any():
                    opciones_indice_item_2.append("ICL INE indice")

            if not opciones_indice_item_2:
                st.warning(
                    "No hay columnas IR INE indice ni ICL INE indice con datos "
                    "para calcular Ítem 2."
                )
            else:
                columna_indice_item_2 = st.selectbox(
                    "Índice para calcular Ítem 2",
                    options=opciones_indice_item_2,
                    index=0
                )

                df_consolidado_calculado = calcular_item_2_item_3(
                    df_consolidado_resumido,
                    columna_indice_item_2=columna_indice_item_2
                )

                st.subheader("Consolidado resumido con columnas calculadas")

                st.dataframe(
                    estilo_columnas_calculadas(df_consolidado_calculado),
                    use_container_width=True
                )

                st.caption(
                    "Las columnas calculadas se muestran destacadas en amarillo: "
                    "Item 2 calculado e Item 3 calculado."
                )

        columnas_graficables = [
            col for col in df_consolidado.columns
            if col not in ["Fecha", "Año", "Mes", "Dolar ultimo dia observado"]
            and pd.api.types.is_numeric_dtype(df_consolidado[col])
        ]

        if calcular_items and df_consolidado_calculado is not None:
            columnas_calculadas_graficables = [
                col for col in ["Item 2 calculado", "Item 3 calculado"]
                if col in df_consolidado_calculado.columns
            ]

            columnas_graficables = columnas_graficables + columnas_calculadas_graficables

        if columnas_graficables:
            st.subheader("Gráfico temporal")

            columnas_seleccionadas = st.multiselect(
                "Selecciona indicadores para graficar",
                options=columnas_graficables,
                default=columnas_graficables[:1]
            )

            if columnas_seleccionadas:
                if calcular_items and df_consolidado_calculado is not None:
                    df_para_grafico = df_consolidado_calculado.copy()
                else:
                    df_para_grafico = df_consolidado.copy()

                columnas_existentes = [
                    col for col in columnas_seleccionadas
                    if col in df_para_grafico.columns
                ]

                if columnas_existentes:
                    df_grafico = df_para_grafico.set_index("Fecha")[columnas_existentes]
                    st.line_chart(df_grafico)
                else:
                    st.info("Las columnas seleccionadas no están disponibles para graficar.")
            else:
                st.info("Selecciona al menos un indicador para graficar.")

        col_descarga_1, col_descarga_2, col_descarga_3 = st.columns(3)

        with col_descarga_1:
            excel_resumido = crear_excel_consolidado_resumido(
                df_consolidado_resumido
            )

            st.download_button(
                label="Descargar Excel resumido",
                data=excel_resumido,
                file_name="consolidado_temporal_resumido.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with col_descarga_2:
            excel_completo = crear_excel_consolidado(
                df_consolidado,
                df_consolidado_resumido
            )

            st.download_button(
                label="Descargar Excel completo",
                data=excel_completo,
                file_name="consolidado_temporal_completo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with col_descarga_3:
            if df_consolidado_calculado is not None:
                excel_calculado = crear_excel_consolidado_calculado(
                    df_consolidado_calculado
                )

                st.download_button(
                    label="Descargar Excel calculado",
                    data=excel_calculado,
                    file_name="consolidado_temporal_calculado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.button(
                    "Descargar Excel calculado",
                    disabled=True
                )
